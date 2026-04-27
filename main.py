import asyncio
import csv
import hashlib
import logging
import os
import re
import sqlite3
import time
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import FSInputFile, KeyboardButton, Message, ReplyKeyboardMarkup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference


# =========================
# CONFIG
# =========================
BOT_TOKEN = os.getenv("BOT_TOKEN")
OWNER_ID = int(os.getenv("OWNER_ID", "0"))

TIMEZONE = ZoneInfo("Europe/Moscow")
DB_PATH = "bot.db"
LOG_PATH = "bot.log"

REMINDER_MINUTES = 10
DOUBLE_TAP_SECONDS = 1.2
SHIFT_PAGE_SIZE = 5
BET_PAGE_SIZE = 20
FAST_AMOUNTS = [1000, 1500, 2000]

# Источник присылает время матча на 1 час позже МСК.
SOURCE_TIME_AHEAD_HOURS = 1

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN not found in environment")
if not OWNER_ID:
    raise RuntimeError("OWNER_ID not found in environment")


# =========================
# LOGGING / BOT
# =========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.FileHandler(LOG_PATH, encoding="utf-8"), logging.StreamHandler()],
)
logger = logging.getLogger("value-bot")

bot = Bot(BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())
scheduler = AsyncIOScheduler(timezone="Europe/Moscow")
ACTION_GUARD = {}


# =========================
# STATES
# =========================
class ShiftState(StatesGroup):
    waiting_budget = State()
    waiting_forwarded_bet = State()
    waiting_risk_decision = State()
    waiting_bet_amount = State()
    waiting_end_shift_confirm = State()
    waiting_shift_number = State()
    waiting_shift_bet_number = State()
    waiting_selected_bet_action = State()
    waiting_selected_bet_new_stake = State()
    waiting_selected_bet_result_status = State()


# =========================
# CONSTANTS
# =========================
RESULT_MAP = {
    "🕒 В ожидании": "pending",
    "✅ Выигрыш": "win",
    "❌ Проигрыш": "lose",
    "🟡 Половина выигрыша": "half_win",
    "🟠 Половина проигрыша": "half_lose",
    "↩️ Возврат": "refund",
}

RESULT_LABELS = {
    "pending": "🕒 В ожидании",
    "win": "✅ Выигрыш",
    "lose": "❌ Проигрыш",
    "half_win": "🟡 Половина выигрыша",
    "half_lose": "🟠 Половина проигрыша",
    "refund": "↩️ Возврат",
}

BOOKMAKER_EMOJI = {
    "fonbet": "🔴",
    "фонбет": "🔴",
    "betcity": "🔵",
    "бетсити": "🔵",
    "betboom": "🟣",
    "бетбум": "🟣",
    "marathon": "🟠",
    "марафон": "🟠",
    "ligastavok": "🟢",
    "лига ставок": "🟢",
    "liga stavok": "🟢",
}


# =========================
# HELPERS
# =========================
def now_dt() -> datetime:
    return datetime.now(TIMEZONE)


def now_str() -> str:
    return now_dt().strftime("%Y-%m-%d %H:%M:%S")


def as_float(text: str) -> float:
    return float((text or "").replace(" ", "").replace(",", "."))


def connect():
    return sqlite3.connect(DB_PATH, check_same_thread=False)


def has_recent_action(user_id: int, action: str, seconds: float = DOUBLE_TAP_SECONDS) -> bool:
    key = f"{user_id}:{action}"
    ts = time.time()
    last = ACTION_GUARD.get(key)
    ACTION_GUARD[key] = ts
    return last is not None and (ts - last) < seconds


def normalize_text(text: str) -> str:
    text = (text or "").lower().replace("ё", "е").replace("−", "-").replace("–", "-")
    text = re.sub(r"[^a-zа-я0-9+\-. ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_html(text: str) -> str:
    text = re.sub(r"</?b>", "", text or "")
    return text.replace("━━━━━━━━━━━━━━", "").strip()


def hash_text(text: str) -> str:
    return hashlib.md5((text or "").strip().encode("utf-8")).hexdigest()


def bookmaker_label(bookmaker: str) -> str:
    norm = normalize_text(bookmaker)
    for key, emoji in BOOKMAKER_EMOJI.items():
        if key in norm:
            return f"{emoji} {bookmaker}"
    return f"⚪ {bookmaker or 'Other'}"


def split_sport_tournament(header: str):
    parts = [x.strip() for x in re.split(r"\s+-\s+", header.strip()) if x.strip()]
    return (parts[0] if parts else ""), (" - ".join(parts[1:]) if len(parts) > 1 else "")


def split_teams(match_name: str):
    parts = re.split(r"\s+[–-]\s+", match_name or "")
    return (parts[0].strip(), parts[1].strip()) if len(parts) >= 2 else ("", "")


def parse_match_start(match_date: str, match_time: str) -> datetime:
    day, month = match_date.split("/")
    hour, minute = match_time.split(":")
    current = now_dt()
    dt = datetime(current.year, int(month), int(day), int(hour), int(minute), tzinfo=TIMEZONE)
    dt -= timedelta(hours=SOURCE_TIME_AHEAD_HOURS)

    if dt < current - timedelta(days=30):
        dt = dt.replace(year=current.year + 1)

    return dt


def event_key_from(match_name: str, match_start_at: str) -> str:
    return f"{normalize_text(match_name)}__{match_start_at[:16]}"


def token_overlap_score(phrase: str, team: str) -> int:
    p = {t for t in normalize_text(phrase).split() if len(t) >= 3}
    q = {t for t in normalize_text(team).split() if len(t) >= 3}
    return len(p & q)


def infer_selection_side(selection: str, team_a: str, team_b: str) -> str:
    s = normalize_text(selection)

    if re.search(r"\bп1\b|\b1\b", s):
        return "home"
    if re.search(r"\bп2\b|\b2\b", s):
        return "away"

    score_a = token_overlap_score(selection, team_a)
    score_b = token_overlap_score(selection, team_b)

    if score_a > score_b and score_a > 0:
        return "home"
    if score_b > score_a and score_b > 0:
        return "away"

    return f"selection:{normalize_text(selection)[:60]}"


def extract_selection_phrase(market: str) -> str:
    m = re.search(r"(.+?)\s+(?:победит\s+с\s+форой|с\s+форой|фора|гандикап)", market, re.I)
    if m:
        return m.group(1).strip(" .,-")

    m = re.search(r"победа\s+(.+?)(?:\s+с\s+учетом|\s+-|\.|$)", market, re.I)
    if m:
        return m.group(1).strip(" .,-")

    return (market or "")[:80]


def parse_line_value(text: str):
    text = (text or "").replace("−", "-").replace(",", ".")

    m = re.search(r"(?:фор[ао]й?|гандикап)[^+\-0-9]{0,25}([+\-]?\d+(?:\.\d+)?)", text, re.I)
    if m:
        return float(m.group(1))

    m = re.search(r"тотал\s+(?:больше|меньше)\s+(\d+(?:\.\d+)?)", text, re.I)
    if m:
        return float(m.group(1))

    return None


def normalize_market(parsed: dict) -> dict:
    market = parsed["market"]
    market_norm = normalize_text(market)
    team_a, team_b = split_teams(parsed["match_name"])
    line_value = parse_line_value(market)

    market_type = "unknown"
    market_side = "unknown"
    market_group = "Другое"
    selection_name = ""
    selection_side = "unknown"
    period_type = "full_match"

    if "1-й сет" in market_norm or "1 сет" in market_norm:
        period_type = "first_set"
    elif "1-й период" in market_norm or "1 период" in market_norm:
        period_type = "first_period"
    elif "с учетом овертайма" in market_norm or "овертаим" in market_norm or "овертайм" in market_norm:
        period_type = "full_match_overtime"

    if "фор" in market_norm or "гандикап" in market_norm:
        market_type = "handicap"
        if line_value is not None and line_value > 0:
            market_side, market_group = "plus", "Фора плюсовая"
        elif line_value is not None and line_value < 0:
            market_side, market_group = "minus", "Фора минусовая"
        else:
            market_side, market_group = "zero", "Фора 0"

        selection_name = extract_selection_phrase(market)
        selection_side = infer_selection_side(selection_name, team_a, team_b)

    elif "тотал больше" in market_norm or re.search(r"\bтб\b", market_norm):
        market_type, market_side, market_group = "total", "over", "Тотал больше"
        selection_name, selection_side = "total", "total"

    elif "тотал меньше" in market_norm or re.search(r"\bтм\b", market_norm):
        market_type, market_side, market_group = "total", "under", "Тотал меньше"
        selection_name, selection_side = "total", "total"

    elif "победа" in market_norm or "победит" in market_norm:
        market_type, market_side, market_group = "moneyline", "win", "Победа"
        selection_name = extract_selection_phrase(market)
        selection_side = infer_selection_side(selection_name, team_a, team_b)

    parsed.update({
        "team_a": team_a,
        "team_b": team_b,
        "selection_name": selection_name,
        "selection_side": selection_side,
        "market_type": market_type,
        "market_group": market_group,
        "market_side": market_side,
        "line_value": line_value,
        "period_type": period_type,
        "semantic_key": "__".join([
            parsed.get("event_key", ""),
            market_type,
            selection_side,
            market_side,
            period_type,
        ]),
    })
    return parsed


def calc_settlement(stake: float, odds: float, result_status: str):
    if result_status == "pending":
        return None, None
    if result_status == "win":
        payout = round(stake * odds, 2)
        return payout, round(payout - stake, 2)
    if result_status == "lose":
        return 0.0, round(-stake, 2)
    if result_status == "half_win":
        payout = round((stake / 2) * odds + (stake / 2), 2)
        return payout, round(payout - stake, 2)
    if result_status == "half_lose":
        payout = round(stake / 2, 2)
        return payout, round(payout - stake, 2)
    if result_status == "refund":
        return round(stake, 2), 0.0
    return None, None


def calc_roi(total_profit: float, total_stake: float) -> float:
    return 0.0 if not total_stake else round((total_profit / total_stake) * 100, 2)


def hedge_amount(stake1: float, odds1: float, odds2: float) -> float:
    return 0.0 if not odds2 else round((stake1 * odds1) / odds2, 2)


def arbitrage_metrics(stake1: float, odds1: float, odds2: float):
    stake2 = hedge_amount(stake1, odds1, odds2)
    payout = round(stake1 * odds1, 2)
    total_stake = round(stake1 + stake2, 2)
    profit = round(payout - total_stake, 2)
    roi = calc_roi(profit, total_stake)
    implied_sum = round(1 / odds1 + 1 / odds2, 4) if odds1 and odds2 else 0
    return stake2, payout, total_stake, profit, roi, implied_sum


def handicap_bound(side: str, line: float):
    # Условие по марже первой команды:
    # home handicap X => margin > -X
    # away handicap X => margin < X
    if side == "home":
        return "gt", -line
    if side == "away":
        return "lt", line
    return None, None


def split_asian_line(line: float) -> list[float]:
    """
    Четвертные азиатские форы делятся на две половины:
    +0.25 = [0, +0.5], -0.25 = [0, -0.5],
    +0.75 = [+0.5, +1.0], -0.75 = [-0.5, -1.0].
    """
    line = round(float(line), 2)
    frac = round(abs(line) % 1, 2)

    if frac in {0.25, 0.75}:
        if line > 0:
            return [round(line - 0.25, 2), round(line + 0.25, 2)]
        return [round(line + 0.25, 2), round(line - 0.25, 2)]

    return [line]


def asian_handicap_profit(selection_side: str, line: float, stake: float, odds: float, margin_home: float) -> float:
    """
    Profit по азиатской форе с учётом четвертных линий.
    margin_home = Team A - Team B.
    home line: margin_home + line
    away line: -margin_home + line
    """
    splits = split_asian_line(line)
    part = stake / len(splits)
    profit = 0.0

    for sub_line in splits:
        result = (margin_home + sub_line) if selection_side == "home" else (-margin_home + sub_line)

        if result > 1e-9:
            profit += part * (odds - 1)
        elif result < -1e-9:
            profit -= part
        else:
            profit += 0.0

    return round(profit, 2)


def corridor_sample_margin(lower: float, upper: float) -> float:
    """
    Выбираем точку внутри коридора для оценки прибыли.
    Приоритет — целая маржа внутри диапазона, например ничья = 0.
    """
    for candidate in range(int(lower) - 5, int(upper) + 6):
        if lower < candidate < upper:
            return float(candidate)
    return (lower + upper) / 2


# =========================
# LOGS
# =========================
def save_log(level: str, message: str):
    db = connect()
    db.execute("""
        CREATE TABLE IF NOT EXISTS logs(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            level TEXT NOT NULL,
            message TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    db.execute(
        "INSERT INTO logs(level, message, created_at) VALUES (?, ?, ?)",
        (level, message, now_str()),
    )
    db.commit()
    db.close()


def log_info(text: str):
    logger.info(text)
    save_log("INFO", text)


def log_warning(text: str):
    logger.warning(text)
    save_log("WARNING", text)


def log_error(text: str):
    logger.error(text)
    save_log("ERROR", text)


# =========================
# DB INIT / MIGRATIONS
# =========================
def add_column_if_not_exists(table_name: str, column_name: str, ddl: str):
    db = connect()
    cols = db.execute(f"PRAGMA table_info({table_name})").fetchall()
    if column_name not in [c[1] for c in cols]:
        db.execute(ddl)
        db.commit()
    db.close()


def init_db():
    db = connect()

    db.execute("""
        CREATE TABLE IF NOT EXISTS shifts(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            started_at TEXT NOT NULL,
            ended_at TEXT,
            budget REAL NOT NULL,
            spent REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS bets(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shift_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            sport TEXT,
            tournament TEXT,
            match_name TEXT,
            match_date TEXT,
            match_time TEXT,
            match_start_at TEXT,
            market TEXT,
            odds REAL,
            ev REAL,
            bookmaker TEXT,
            stake REAL,
            source_text TEXT,
            match_hash TEXT UNIQUE,
            reminder_sent INTEGER DEFAULT 0,
            result_status TEXT DEFAULT 'pending',
            payout REAL,
            profit REAL,
            event_key TEXT,
            team_a TEXT,
            team_b TEXT,
            selection_name TEXT,
            selection_side TEXT,
            market_type TEXT,
            market_group TEXT,
            market_side TEXT,
            line_value REAL,
            period_type TEXT,
            semantic_key TEXT,
            risk_status TEXT DEFAULT 'new',
            risk_notes TEXT
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS rejected_bets(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            sport TEXT,
            tournament TEXT,
            match_name TEXT,
            market TEXT,
            odds REAL,
            ev REAL,
            bookmaker TEXT,
            source_text TEXT,
            event_key TEXT,
            market_type TEXT,
            market_group TEXT,
            risk_status TEXT,
            risk_notes TEXT,
            reason TEXT
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS logs(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            level TEXT NOT NULL,
            message TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)

    db.commit()
    db.close()

    migrations = [
        ("bets", "match_start_at", "ALTER TABLE bets ADD COLUMN match_start_at TEXT"),
        ("bets", "reminder_sent", "ALTER TABLE bets ADD COLUMN reminder_sent INTEGER DEFAULT 0"),
        ("bets", "result_status", "ALTER TABLE bets ADD COLUMN result_status TEXT DEFAULT 'pending'"),
        ("bets", "payout", "ALTER TABLE bets ADD COLUMN payout REAL"),
        ("bets", "profit", "ALTER TABLE bets ADD COLUMN profit REAL"),
        ("bets", "event_key", "ALTER TABLE bets ADD COLUMN event_key TEXT"),
        ("bets", "team_a", "ALTER TABLE bets ADD COLUMN team_a TEXT"),
        ("bets", "team_b", "ALTER TABLE bets ADD COLUMN team_b TEXT"),
        ("bets", "selection_name", "ALTER TABLE bets ADD COLUMN selection_name TEXT"),
        ("bets", "selection_side", "ALTER TABLE bets ADD COLUMN selection_side TEXT"),
        ("bets", "market_type", "ALTER TABLE bets ADD COLUMN market_type TEXT"),
        ("bets", "market_group", "ALTER TABLE bets ADD COLUMN market_group TEXT"),
        ("bets", "market_side", "ALTER TABLE bets ADD COLUMN market_side TEXT"),
        ("bets", "line_value", "ALTER TABLE bets ADD COLUMN line_value REAL"),
        ("bets", "period_type", "ALTER TABLE bets ADD COLUMN period_type TEXT"),
        ("bets", "semantic_key", "ALTER TABLE bets ADD COLUMN semantic_key TEXT"),
        ("bets", "risk_status", "ALTER TABLE bets ADD COLUMN risk_status TEXT DEFAULT 'new'"),
        ("bets", "risk_notes", "ALTER TABLE bets ADD COLUMN risk_notes TEXT"),
    ]

    for table, column, ddl in migrations:
        add_column_if_not_exists(table, column, ddl)

    db = connect()
    db.execute("CREATE INDEX IF NOT EXISTS idx_bets_event_pending ON bets(event_key, result_status)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_bets_shift ON bets(shift_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_bets_user_created ON bets(user_id, created_at)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_bets_match_start ON bets(match_start_at, reminder_sent, result_status)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_bets_risk_status ON bets(risk_status)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_rejected_user_risk ON rejected_bets(user_id, risk_status)")
    db.commit()
    db.close()

    log_info("Database initialized")


# =========================
# DB OPERATIONS
# =========================
def get_active_shift(user_id: int):
    db = connect()
    row = db.execute("""
        SELECT id, budget, spent, started_at
        FROM shifts
        WHERE user_id = ? AND status = 'active'
        ORDER BY id DESC
        LIMIT 1
    """, (user_id,)).fetchone()
    db.close()
    return row


def start_shift_db(user_id: int, started_at: str, budget: float):
    db = connect()
    db.execute(
        "INSERT INTO shifts(user_id, started_at, budget, spent, status) VALUES (?, ?, ?, 0, 'active')",
        (user_id, started_at, budget),
    )
    db.commit()
    db.close()
    log_info(f"Shift started | user={user_id} | budget={budget}")


def end_shift_db(shift_id: int, ended_at: str):
    db = connect()
    db.execute("UPDATE shifts SET ended_at = ?, status = 'ended' WHERE id = ?", (ended_at, shift_id))
    db.commit()
    db.close()
    log_info(f"Shift ended | shift_id={shift_id}")


def add_bet_db(shift_id: int, user_id: int, created_at: str, parsed: dict, stake: float):
    db = connect()
    db.execute("""
        INSERT INTO bets(
            shift_id, user_id, created_at,
            sport, tournament, match_name, match_date, match_time, match_start_at,
            market, odds, ev, bookmaker, stake, source_text, match_hash,
            reminder_sent, result_status, payout, profit,
            event_key, team_a, team_b, selection_name, selection_side,
            market_type, market_group, market_side, line_value, period_type,
            semantic_key, risk_status, risk_notes
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                0, 'pending', NULL, NULL,
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        shift_id,
        user_id,
        created_at,
        parsed["sport"],
        parsed["tournament"],
        parsed["match_name"],
        parsed["match_date"],
        parsed["match_time"],
        parsed["match_start_at"],
        parsed["market"],
        parsed["odds"],
        parsed["ev"],
        parsed["bookmaker"],
        stake,
        parsed["source_text"],
        parsed["hash"],
        parsed.get("event_key"),
        parsed.get("team_a"),
        parsed.get("team_b"),
        parsed.get("selection_name"),
        parsed.get("selection_side"),
        parsed.get("market_type"),
        parsed.get("market_group"),
        parsed.get("market_side"),
        parsed.get("line_value"),
        parsed.get("period_type"),
        parsed.get("semantic_key"),
        parsed.get("risk_status", "new"),
        clean_html(parsed.get("risk_notes", "")),
    ))
    db.execute("UPDATE shifts SET spent = spent + ? WHERE id = ?", (stake, shift_id))
    db.commit()
    db.close()
    log_info(f"Bet added | shift_id={shift_id} | user={user_id} | stake={stake} | risk={parsed.get('risk_status')}")


def save_rejected_bet(user_id: int, parsed: dict, reason: str):
    db = connect()
    db.execute("""
        INSERT INTO rejected_bets(
            created_at, user_id, sport, tournament, match_name, market,
            odds, ev, bookmaker, source_text, event_key, market_type,
            market_group, risk_status, risk_notes, reason
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        now_str(),
        user_id,
        parsed.get("sport"),
        parsed.get("tournament"),
        parsed.get("match_name"),
        parsed.get("market"),
        parsed.get("odds"),
        parsed.get("ev"),
        parsed.get("bookmaker"),
        parsed.get("source_text"),
        parsed.get("event_key"),
        parsed.get("market_type"),
        parsed.get("market_group"),
        parsed.get("risk_status"),
        clean_html(parsed.get("risk_notes")),
        reason,
    ))
    db.commit()
    db.close()
    log_warning(f"Rejected bet saved | reason={reason} | risk={parsed.get('risk_status')}")


def get_pending_similar_bets(parsed: dict):
    db = connect()
    rows = db.execute("""
        SELECT id, sport, match_name, market, odds, stake, bookmaker, result_status,
               selection_side, market_type, market_side, line_value, event_key,
               semantic_key, market_group
        FROM bets
        WHERE event_key = ? AND result_status = 'pending'
        ORDER BY id DESC
    """, (parsed.get("event_key"),)).fetchall()
    db.close()
    return rows


def get_last_bets(user_id: int, limit: int = 20):
    db = connect()
    rows = db.execute("""
        SELECT id, sport, match_name, market, odds, stake, bookmaker, created_at, result_status
        FROM bets
        WHERE user_id = ?
        ORDER BY id DESC
        LIMIT ?
    """, (user_id, limit)).fetchall()
    db.close()
    return rows


def get_last_bet(user_id: int):
    rows = get_last_bets(user_id, 1)
    return rows[0] if rows else None


def get_bet_by_id(bet_id: int):
    db = connect()
    row = db.execute("""
        SELECT id, sport, tournament, match_name, match_date, match_time,
               market, odds, ev, bookmaker, stake, created_at, result_status,
               payout, profit, shift_id
        FROM bets
        WHERE id = ?
    """, (bet_id,)).fetchone()
    db.close()
    return row


def get_shift_by_id(shift_id: int, user_id: int):
    db = connect()
    row = db.execute("""
        SELECT id, user_id, started_at, ended_at, budget, spent, status
        FROM shifts
        WHERE id = ? AND user_id = ?
    """, (shift_id, user_id)).fetchone()
    db.close()
    return row


def list_shifts(user_id: int, offset: int = 0, limit: int = SHIFT_PAGE_SIZE):
    db = connect()
    rows = db.execute("""
        SELECT id, started_at, ended_at, budget, spent, status
        FROM shifts
        WHERE user_id = ?
        ORDER BY id DESC
        LIMIT ? OFFSET ?
    """, (user_id, limit, offset)).fetchall()
    total = db.execute("SELECT COUNT(*) FROM shifts WHERE user_id = ?", (user_id,)).fetchone()[0]
    db.close()
    return rows, total


def get_bets_by_shift(shift_id: int, offset: int = 0, limit: int = BET_PAGE_SIZE):
    db = connect()
    rows = db.execute("""
        SELECT id, sport, match_name, market, odds, stake, bookmaker, created_at, result_status
        FROM bets
        WHERE shift_id = ?
        ORDER BY
            CASE WHEN result_status='pending' THEN 0 ELSE 1 END,
            COALESCE(match_start_at, created_at) ASC
        LIMIT ? OFFSET ?
    """, (shift_id, limit, offset)).fetchall()
    total = db.execute("SELECT COUNT(*) FROM bets WHERE shift_id = ?", (shift_id,)).fetchone()[0]
    db.close()
    return rows, total


def get_shift_stats(shift_id: int):
    db = connect()
    row = db.execute("""
        SELECT
            COUNT(*),
            COALESCE(SUM(stake), 0),
            COALESCE(AVG(odds), 0),
            COALESCE(AVG(ev), 0),
            COALESCE(SUM(CASE WHEN result_status='win' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='lose' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='half_win' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='half_lose' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='refund' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='pending' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(profit), 0)
        FROM bets
        WHERE shift_id = ?
    """, (shift_id,)).fetchone()
    db.close()
    return row


def get_global_stats(user_id: int):
    db = connect()
    row = db.execute("""
        SELECT
            COUNT(*),
            COALESCE(SUM(stake), 0),
            COALESCE(AVG(odds), 0),
            COALESCE(AVG(ev), 0),
            COALESCE(SUM(CASE WHEN result_status='win' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='lose' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='half_win' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='half_lose' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='refund' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='pending' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(profit), 0)
        FROM bets
        WHERE user_id = ?
    """, (user_id,)).fetchone()
    db.close()
    return row


def get_market_stats_by_shift(shift_id: int):
    db = connect()
    rows = db.execute("""
        SELECT
            COALESCE(market_group, 'Другое'),
            COUNT(*),
            COALESCE(SUM(stake), 0),
            COALESCE(AVG(stake), 0),
            COALESCE(AVG(odds), 0),
            COALESCE(AVG(ev), 0),
            COALESCE(SUM(profit), 0),
            COALESCE(SUM(CASE WHEN result_status='pending' THEN 1 ELSE 0 END), 0)
        FROM bets
        WHERE shift_id = ?
        GROUP BY COALESCE(market_group, 'Другое')
        ORDER BY COUNT(*) DESC
    """, (shift_id,)).fetchall()
    total = db.execute("SELECT COUNT(*) FROM bets WHERE shift_id = ?", (shift_id,)).fetchone()[0]
    db.close()
    return rows, total


def get_bookmaker_stats_by_shift(shift_id: int):
    db = connect()
    rows = db.execute("""
        SELECT
            COALESCE(bookmaker, 'Other'),
            COUNT(*),
            COALESCE(SUM(stake), 0),
            COALESCE(AVG(stake), 0),
            COALESCE(AVG(odds), 0),
            COALESCE(SUM(profit), 0)
        FROM bets
        WHERE shift_id = ?
        GROUP BY COALESCE(bookmaker, 'Other')
        ORDER BY COUNT(*) DESC
    """, (shift_id,)).fetchall()
    total = db.execute("SELECT COUNT(*) FROM bets WHERE shift_id = ?", (shift_id,)).fetchone()[0]
    db.close()
    return rows, total


def get_risk_stats_by_shift(shift_id: int):
    db = connect()
    rows = db.execute("""
        SELECT
            COALESCE(risk_status, 'new'),
            COUNT(*),
            COALESCE(SUM(stake), 0),
            COALESCE(AVG(stake), 0),
            COALESCE(AVG(odds), 0),
            COALESCE(AVG(ev), 0),
            COALESCE(SUM(profit), 0),
            COALESCE(SUM(CASE WHEN result_status='pending' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='win' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='lose' THEN 1 ELSE 0 END), 0)
        FROM bets
        WHERE shift_id = ?
        GROUP BY COALESCE(risk_status, 'new')
        ORDER BY COUNT(*) DESC
    """, (shift_id,)).fetchall()
    total = db.execute("SELECT COUNT(*) FROM bets WHERE shift_id = ?", (shift_id,)).fetchone()[0]
    db.close()
    return rows, total


def get_risk_stats_all(user_id: int):
    db = connect()
    saved = db.execute("""
        SELECT
            COALESCE(risk_status, 'new'),
            COUNT(*),
            COALESCE(SUM(stake), 0),
            COALESCE(AVG(stake), 0),
            COALESCE(AVG(odds), 0),
            COALESCE(AVG(ev), 0),
            COALESCE(SUM(profit), 0),
            COALESCE(SUM(CASE WHEN result_status='pending' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='win' THEN 1 ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN result_status='lose' THEN 1 ELSE 0 END), 0)
        FROM bets
        WHERE user_id = ?
        GROUP BY COALESCE(risk_status, 'new')
        ORDER BY COUNT(*) DESC
    """, (user_id,)).fetchall()
    rejected = db.execute("""
        SELECT COALESCE(risk_status, 'unknown'), COUNT(*)
        FROM rejected_bets
        WHERE user_id = ?
        GROUP BY COALESCE(risk_status, 'unknown')
    """, (user_id,)).fetchall()
    total = db.execute("SELECT COUNT(*) FROM bets WHERE user_id = ?", (user_id,)).fetchone()[0]
    db.close()
    return saved, dict(rejected), total


def update_bet_result(bet_id: int, result_status: str):
    db = connect()
    row = db.execute("SELECT stake, odds FROM bets WHERE id = ?", (bet_id,)).fetchone()
    if not row:
        db.close()
        return False

    stake, odds = row
    payout, profit = calc_settlement(stake, odds, result_status)
    db.execute(
        "UPDATE bets SET result_status = ?, payout = ?, profit = ? WHERE id = ?",
        (result_status, payout, profit, bet_id),
    )
    db.commit()
    db.close()
    log_info(f"Bet result updated | bet_id={bet_id} | status={result_status}")
    return True


def update_bet_stake(bet_id: int, new_stake: float):
    db = connect()
    row = db.execute(
        "SELECT shift_id, stake, odds, result_status FROM bets WHERE id = ?",
        (bet_id,),
    ).fetchone()

    if not row:
        db.close()
        return False, "Ставка не найдена."

    shift_id, old_stake, odds, result_status = row
    delta = new_stake - old_stake
    payout, profit = calc_settlement(new_stake, odds, result_status)

    db.execute(
        "UPDATE bets SET stake = ?, payout = ?, profit = ? WHERE id = ?",
        (new_stake, payout, profit, bet_id),
    )
    db.execute("UPDATE shifts SET spent = spent + ? WHERE id = ?", (delta, shift_id))
    db.commit()
    db.close()
    log_info(f"Bet stake updated | bet_id={bet_id} | old={old_stake} | new={new_stake}")
    return True, old_stake


def get_recent_logs(limit: int = 10):
    db = connect()
    rows = db.execute(
        "SELECT level, message, created_at FROM logs ORDER BY id DESC LIMIT ?",
        (limit,),
    ).fetchall()
    db.close()
    return rows


def get_due_reminders():
    db = connect()
    rows = db.execute("""
        SELECT id, user_id, match_name, market, match_start_at, stake, odds, bookmaker
        FROM bets
        WHERE reminder_sent = 0
          AND result_status = 'pending'
          AND match_start_at IS NOT NULL
    """).fetchall()
    db.close()

    current = now_dt()
    upper = current + timedelta(minutes=REMINDER_MINUTES)
    due = []

    for row in rows:
        bet_id, user_id, match_name, market, match_start_at, stake, odds, bookmaker = row
        try:
            dt = datetime.fromisoformat(match_start_at)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=TIMEZONE)
        except Exception:
            continue

        if current <= dt <= upper:
            due.append({
                "id": bet_id,
                "user_id": user_id,
                "match_name": match_name,
                "market": market,
                "match_start_at": dt,
                "stake": stake,
                "odds": odds,
                "bookmaker": bookmaker,
            })

    return due


def mark_reminder_sent(bet_id: int):
    db = connect()
    db.execute("UPDATE bets SET reminder_sent = 1 WHERE id = ?", (bet_id,))
    db.commit()
    db.close()


# =========================
# PARSER
# =========================
def parse_bet_diagnostics(text: str) -> str:
    checks = []

    if "⚽️🏒🎾" not in text:
        checks.append("не нашёл строку спорта/турнира с маркером ⚽️🏒🎾")
    if "🚩" not in text:
        checks.append("не нашёл строку матча с маркером 🚩")
    if "коэф" not in normalize_text(text):
        checks.append("не нашёл коэффициент: в тексте нет слова 'коэф.'")
    if not re.search(r"\d{1,2}:\d{2}\s+\d{2}/\d{2}", text):
        checks.append("не нашёл время и дату матча в формате 12:30 28/03")
    if "Ставка сделана" not in text:
        checks.append("не нашёл букмекера: строку 'Ставка сделана👉 ...'")

    if not checks:
        checks.append("структура похожа на ставку, но формат отличается от ожидаемого шаблона")

    return "\n".join([f"• {item}" for item in checks])


def parse_bet(text: str):
    pattern = re.compile(r"(⚽️🏒🎾.*?)(?=\n\s*⚽️🏒🎾|\Z)", re.S)
    matches = pattern.findall(text)

    if not matches:
        return None

    block = None
    for candidate in matches:
        sport_ok = re.search(r"⚽️🏒🎾\s*(.+?)\n", candidate, re.S)
        event_ok = re.search(r"🚩\s*(.+?),\s*(\d{1,2}:\d{2})\s+(\d{2}/\d{2})", candidate)
        market_ok = re.search(r"❗️\s*(.+?)\s*коэф\.?\s*([\d.,]+)❗️?", candidate, re.S)

        if sport_ok and event_ok and market_ok:
            block = candidate.strip()
            break

    if not block:
        return None

    sport_line = re.search(r"⚽️🏒🎾\s*(.+?)\n", block, re.S)
    event_line = re.search(r"🚩\s*(.+?),\s*(\d{1,2}:\d{2})\s+(\d{2}/\d{2})", block)
    market_line = re.search(r"❗️\s*(.+?)\s*коэф\.?\s*([\d.,]+)❗️?", block, re.S)
    ev_line = re.search(r"Математическое ожидание\s*≈\s*([\d.,]+)%", block, re.I)
    bk_line = re.search(r"Ставка сделана👉\s*([^\n(]+)", block, re.I)

    if not sport_line or not event_line or not market_line:
        return None

    full_header = sport_line.group(1).strip()
    sport, tournament = split_sport_tournament(full_header)

    match_name = event_line.group(1).strip()
    match_time = event_line.group(2).strip()
    match_date = event_line.group(3).strip()
    market = re.sub(r"\s+", " ", market_line.group(1)).strip()

    odds = float(market_line.group(2).replace(",", "."))
    ev = float(ev_line.group(1).replace(",", ".")) if ev_line else None
    bookmaker = bk_line.group(1).strip() if bk_line else ""

    match_start_at = parse_match_start(match_date, match_time)

    parsed = {
        "sport": sport,
        "tournament": tournament,
        "match_name": match_name,
        "match_time": match_time,
        "match_date": match_date,
        "match_start_at": match_start_at.isoformat(),
        "market": market,
        "odds": odds,
        "ev": ev,
        "bookmaker": bookmaker,
        "hash": hash_text(block),
        "source_text": block,
    }

    parsed["event_key"] = event_key_from(match_name, parsed["match_start_at"])
    return normalize_market(parsed)


# =========================
# RISK ENGINE
# =========================
def make_risk_note(kind: str, existing: dict | None = None, parsed: dict | None = None, extra: dict | None = None) -> str:
    extra = extra or {}

    if kind == "new":
        return "NEW | pending-повторов, коридоров и вилок не найдено"

    if kind == "duplicate":
        return (
            f"ПОВТОР | уже: {existing['market']} @ {existing['odds']} / {existing['stake']} / {existing['bookmaker']} | "
            f"новая: {parsed['market']} @ {parsed['odds']} / {parsed['bookmaker']} | та же сторона и тип рынка"
        )

    if kind == "corridor":
        return (
            f"КОРИДОР | уже: {existing['market']} @ {existing['odds']} / {existing['stake']} | "
            f"новая: {parsed['market']} @ {parsed['odds']} | "
            f"диапазон: {extra.get('range')} | ширина: {extra.get('width')} | "
            f"пример маржи: {extra.get('sample_margin')} | реком. сумма: {extra.get('stake2')} | "
            f"в коридоре: {extra.get('corridor_profit')} | вне: {extra.get('outside_profit')} ({extra.get('outside_roi')}%)"
        )

    if kind == "arbitrage":
        return (
            f"ВИЛКА | уже: {existing['market']} @ {existing['odds']} / {existing['stake']} | "
            f"новая: {parsed['market']} @ {parsed['odds']} | "
            f"вероятности: {extra.get('implied')} | реком. сумма: {extra.get('stake2')} | результат: {extra.get('profit')} ({extra.get('roi')}%)"
        )

    if kind == "opposite_no_value":
        return (
            f"ПЛЕЧО БЕЗ ПЛЮСА | уже: {existing['market']} @ {existing['odds']} / {existing['stake']} | "
            f"новая: {parsed['market']} @ {parsed['odds']} | "
            f"вероятности: {extra.get('implied')} | реком. сумма: {extra.get('stake2')} | результат: {extra.get('profit')} ({extra.get('roi')}%)"
        )

    return kind


def risk_note_to_telegram(risk_status: str, risk_notes: str) -> str:
    if risk_status == "new":
        return "✅ <b>ЧИСТО</b>\nPending-повторов / коридоров / вилок не найдено."

    title_map = {
        "duplicate": "⚠️ <b>ПОВТОР</b>",
        "corridor": "🟣 <b>КОРИДОР</b>",
        "arbitrage": "🟢 <b>ВИЛКА</b>",
        "mixed": "🔀 <b>ВИЛКА / КОРИДОР</b>",
        "opposite_no_value": "🟡 <b>ВИЛКИ НЕТ</b>",
    }

    title = title_map.get(risk_status, f"⚠️ <b>{risk_status}</b>")
    parts = risk_notes.split(" | ")
    lines = []

    for part in parts[1:]:
        if ":" in part:
            left, right = part.split(":", 1)
            lines.append(f"<b>{left.strip()}:</b> {right.strip()}")
        else:
            lines.append(part.strip())

    return title + "\n" + "\n".join(lines[:7])


def analyze_risk(parsed: dict):
    rows = get_pending_similar_bets(parsed)
    notes = []
    statuses = []

    for row in rows:
        existing = {
            "id": row[0],
            "sport": row[1],
            "match_name": row[2],
            "market": row[3],
            "odds": row[4],
            "stake": row[5],
            "bookmaker": row[6],
            "selection_side": row[8],
            "market_type": row[9],
            "market_side": row[10],
            "line_value": row[11],
            "event_key": row[12],
            "semantic_key": row[13],
            "market_group": row[14],
        }

        same_type = existing["market_type"] == parsed.get("market_type")
        same_selection = (
            existing["selection_side"] == parsed.get("selection_side")
            and parsed.get("selection_side") not in {None, "unknown"}
        )

        if same_type and same_selection and parsed.get("market_type") in {"handicap", "moneyline"}:
            statuses.append("duplicate")
            notes.append(make_risk_note("duplicate", existing, parsed))
            continue

        if same_type and parsed.get("market_type") == "total" and existing.get("market_side") == parsed.get("market_side"):
            statuses.append("duplicate")
            notes.append(make_risk_note("duplicate", existing, parsed))
            continue

        if (
            same_type
            and parsed.get("market_type") == "moneyline"
            and existing.get("selection_side") in {"home", "away"}
            and parsed.get("selection_side") in {"home", "away"}
            and existing.get("selection_side") != parsed.get("selection_side")
        ):
            stake2, payout, total_stake, hedge_profit, hedge_roi, implied = arbitrage_metrics(
                existing["stake"], existing["odds"], parsed["odds"]
            )
            kind = "arbitrage" if implied < 1 else "opposite_no_value"
            statuses.append(kind)
            notes.append(make_risk_note(kind, existing, parsed, {
                "implied": implied,
                "stake2": stake2,
                "profit": hedge_profit,
                "roi": hedge_roi,
            }))
            continue

        if (
            same_type
            and parsed.get("market_type") == "handicap"
            and existing.get("selection_side") in {"home", "away"}
            and parsed.get("selection_side") in {"home", "away"}
            and existing.get("selection_side") != parsed.get("selection_side")
        ):
            line1 = existing.get("line_value")
            line2 = parsed.get("line_value")
            if line1 is None or line2 is None:
                continue

            cond1, val1 = handicap_bound(existing["selection_side"], float(line1))
            cond2, val2 = handicap_bound(parsed["selection_side"], float(line2))

            if cond1 and cond2:
                lower = max([v for c, v in [(cond1, val1), (cond2, val2)] if c == "gt"], default=None)
                upper = min([v for c, v in [(cond1, val1), (cond2, val2)] if c == "lt"], default=None)

                stake2, payout, total_stake, hedge_profit, hedge_roi, implied = arbitrage_metrics(
                    existing["stake"], existing["odds"], parsed["odds"]
                )

                if lower is not None and upper is not None and lower < upper:
                    width = round(upper - lower, 2)
                    sample_margin = corridor_sample_margin(lower, upper)

                    existing_corridor_profit = asian_handicap_profit(
                        existing["selection_side"],
                        float(line1),
                        existing["stake"],
                        existing["odds"],
                        sample_margin,
                    )
                    new_corridor_profit = asian_handicap_profit(
                        parsed["selection_side"],
                        float(line2),
                        stake2,
                        parsed["odds"],
                        sample_margin,
                    )
                    corridor_profit = round(existing_corridor_profit + new_corridor_profit, 2)

                    statuses.append("corridor")
                    notes.append(make_risk_note("corridor", existing, parsed, {
                        "range": f"{lower}—{upper}",
                        "width": width,
                        "sample_margin": sample_margin,
                        "stake2": stake2,
                        "corridor_profit": corridor_profit,
                        "outside_profit": hedge_profit,
                        "outside_roi": hedge_roi,
                    }))
                else:
                    kind = "arbitrage" if implied < 1 else "opposite_no_value"
                    statuses.append(kind)
                    notes.append(make_risk_note(kind, existing, parsed, {
                        "implied": implied,
                        "stake2": stake2,
                        "profit": hedge_profit,
                        "roi": hedge_roi,
                    }))

        if same_type and parsed.get("market_type") == "total" and existing.get("market_side") != parsed.get("market_side"):
            line1 = existing.get("line_value")
            line2 = parsed.get("line_value")
            if line1 is None or line2 is None:
                continue

            over_line = line1 if existing["market_side"] == "over" else line2 if parsed["market_side"] == "over" else None
            under_line = line1 if existing["market_side"] == "under" else line2 if parsed["market_side"] == "under" else None

            stake2, payout, total_stake, hedge_profit, hedge_roi, implied = arbitrage_metrics(
                existing["stake"], existing["odds"], parsed["odds"]
            )

            if over_line is not None and under_line is not None and over_line < under_line:
                width = round(under_line - over_line, 2)
                corridor_profit = round(existing["stake"] * (existing["odds"] - 1) + stake2 * (parsed["odds"] - 1), 2)
                statuses.append("corridor")
                notes.append(make_risk_note("corridor", existing, parsed, {
                    "range": f"{over_line}—{under_line}",
                    "width": width,
                    "sample_margin": "-",
                    "stake2": stake2,
                    "corridor_profit": corridor_profit,
                    "outside_profit": hedge_profit,
                    "outside_roi": hedge_roi,
                }))
            else:
                kind = "arbitrage" if implied < 1 else "opposite_no_value"
                statuses.append(kind)
                notes.append(make_risk_note(kind, existing, parsed, {
                    "implied": implied,
                    "stake2": stake2,
                    "profit": hedge_profit,
                    "roi": hedge_roi,
                }))

    if not statuses:
        status = "new"
        notes_text = make_risk_note("new")
    elif "duplicate" in statuses:
        status = "duplicate"
        duplicate_notes = [n for n in notes if n.startswith("ПОВТОР")]
        notes_text = "\n".join(duplicate_notes or notes)
    elif "corridor" in statuses and "arbitrage" in statuses:
        status = "mixed"
        notes_text = "\n".join(notes)
    elif "arbitrage" in statuses:
        status = "arbitrage"
        notes_text = "\n".join([n for n in notes if n.startswith("ВИЛКА")] or notes)
    elif "corridor" in statuses:
        status = "corridor"
        notes_text = "\n".join(notes)
    else:
        status = "opposite_no_value"
        notes_text = "\n".join(notes)

    parsed["risk_status"] = status
    parsed["risk_notes"] = notes_text
    return parsed


# =========================
# EXPORTS
# =========================
def export_bets_to_csv(user_id: int) -> str | None:
    db = connect()
    rows = db.execute("""
        SELECT id, shift_id, created_at, sport, tournament, match_name,
               match_date, match_time, match_start_at, market, market_group,
               odds, ev, bookmaker, stake, result_status, payout, profit,
               risk_status, risk_notes
        FROM bets
        WHERE user_id = ?
        ORDER BY id DESC
    """, (user_id,)).fetchall()
    db.close()

    if not rows:
        return None

    filename = f"bets_export_all_{now_dt().strftime('%Y%m%d_%H%M%S')}.csv"
    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([
            "id", "shift_id", "created_at", "sport", "tournament", "match_name",
            "match_date", "match_time", "match_start_at", "market", "market_group",
            "odds", "ev", "bookmaker", "stake", "result_status", "payout", "profit",
            "risk_status", "risk_notes"
        ])
        writer.writerows(rows)

    return filename


def style_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    positive_fill = PatternFill("solid", fgColor="E2F0D9")
    negative_fill = PatternFill("solid", fgColor="FCE4D6")
    neutral_fill = PatternFill("solid", fgColor="FFF2CC")

    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill

    for row in sheet.iter_rows():
        if len(row) >= 2:
            label = str(row[0].value or "").lower()
            value_cell = row[1]
            if label in {"profit", "roi %"}:
                try:
                    value = float(value_cell.value or 0)
                    value_cell.fill = positive_fill if value >= 0 else negative_fill
                    value_cell.font = Font(bold=True)
                except Exception:
                    pass
            elif label in {"pending", "bets", "total bets"}:
                value_cell.fill = neutral_fill
                value_cell.font = Font(bold=True)

    for column_cells in sheet.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 55)


def add_summary_charts(sheet):
    try:
        labels = [cell.value for cell in sheet["A"]]
        status_names = {"Win", "Lose", "Half win", "Half lose", "Refund", "Pending"}
        rows = [i + 1 for i, label in enumerate(labels) if label in status_names]

        if rows:
            pie = PieChart()
            pie.title = "Распределение результатов"
            data = Reference(sheet, min_col=2, min_row=min(rows), max_row=max(rows))
            cats = Reference(sheet, min_col=1, min_row=min(rows), max_row=max(rows))
            pie.add_data(data, titles_from_data=False)
            pie.set_categories(cats)
            pie.height = 7
            pie.width = 9
            sheet.add_chart(pie, "D2")

        market_header = None
        bookmaker_header = None
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == "Market" and sheet.cell(row=row, column=2).value == "Count":
                market_header = row
            if sheet.cell(row=row, column=1).value == "Bookmaker" and sheet.cell(row=row, column=2).value == "Count":
                bookmaker_header = row

        if market_header:
            last = market_header
            while last + 1 <= sheet.max_row and sheet.cell(row=last + 1, column=1).value:
                last += 1
            if last > market_header:
                bar = BarChart()
                bar.type = "bar"
                bar.title = "Ставки по маркетам"
                data = Reference(sheet, min_col=2, min_row=market_header, max_row=last)
                cats = Reference(sheet, min_col=1, min_row=market_header + 1, max_row=last)
                bar.add_data(data, titles_from_data=True)
                bar.set_categories(cats)
                bar.height = 8
                bar.width = 13
                sheet.add_chart(bar, "D18")

        if bookmaker_header:
            last = bookmaker_header
            while last + 1 <= sheet.max_row and sheet.cell(row=last + 1, column=1).value:
                last += 1
            if last > bookmaker_header:
                bar2 = BarChart()
                bar2.type = "bar"
                bar2.title = "Ставки по букмекерам"
                data = Reference(sheet, min_col=2, min_row=bookmaker_header, max_row=last)
                cats = Reference(sheet, min_col=1, min_row=bookmaker_header + 1, max_row=last)
                bar2.add_data(data, titles_from_data=True)
                bar2.set_categories(cats)
                bar2.height = 8
                bar2.width = 13
                sheet.add_chart(bar2, "D35")
    except Exception as e:
        logger.warning(f"Excel chart creation failed: {e}")


def export_all_to_xlsx(user_id: int) -> str | None:
    db = connect()
    bets = db.execute("""
        SELECT id, shift_id, created_at, sport, tournament, match_name, match_date,
               match_time, match_start_at, market, market_group, odds, ev,
               bookmaker, stake, result_status, payout, profit, risk_status, risk_notes
        FROM bets
        WHERE user_id = ?
        ORDER BY id ASC
    """, (user_id,)).fetchall()

    shifts = db.execute("""
        SELECT id, started_at, ended_at, budget, spent, status
        FROM shifts
        WHERE user_id = ?
        ORDER BY id DESC
    """, (user_id,)).fetchall()

    rejected = db.execute("""
        SELECT created_at, match_name, market, odds, bookmaker, risk_status, risk_notes, reason
        FROM rejected_bets
        WHERE user_id = ?
        ORDER BY id DESC
    """, (user_id,)).fetchall()
    db.close()

    if not bets and not shifts:
        return None

    filename = f"all_report_{now_dt().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    total_bets, total_stake, avg_odds, avg_ev, wins, loses, half_wins, half_loses, refunds, pendings, total_profit = get_global_stats(user_id)
    ws.append(["Metric", "Value"])
    for row in [
        ["Total bets", total_bets],
        ["Total stake", round(total_stake, 2)],
        ["Avg odds", round(avg_odds, 2)],
        ["Avg EV", round(avg_ev, 2)],
        ["Profit", round(total_profit, 2)],
        ["ROI %", calc_roi(total_profit, total_stake)],
        ["Win", wins],
        ["Lose", loses],
        ["Half win", half_wins],
        ["Half lose", half_loses],
        ["Refund", refunds],
        ["Pending", pendings],
    ]:
        ws.append(row)

    ws.append([])
    ws.append(["Risk status", "Saved bets", "Rejected bets", "Total stake", "Avg stake", "Avg odds", "Avg EV", "Profit", "ROI %", "Pending", "Win", "Lose"])
    risk_rows, rejected_map, risk_total = get_risk_stats_all(user_id)
    for risk_status, count, total_stake_r, avg_stake, avg_odds_r, avg_ev_r, profit_r, pending_r, win_r, lose_r in risk_rows:
        ws.append([
            risk_status,
            count,
            rejected_map.get(risk_status, 0),
            round(total_stake_r, 2),
            round(avg_stake, 2),
            round(avg_odds_r, 2),
            round(avg_ev_r, 2),
            round(profit_r, 2),
            calc_roi(profit_r, total_stake_r),
            pending_r,
            win_r,
            lose_r,
        ])

    add_summary_charts(ws)
    style_sheet(ws)

    ws_risk = wb.create_sheet("Risk_Analytics")
    ws_risk.append(["Risk status", "Saved bets", "Rejected bets", "Total stake", "Avg stake", "Avg odds", "Avg EV", "Profit", "ROI %", "Pending", "Win", "Lose"])
    for risk_status, count, total_stake_r, avg_stake, avg_odds_r, avg_ev_r, profit_r, pending_r, win_r, lose_r in risk_rows:
        ws_risk.append([
            risk_status,
            count,
            rejected_map.get(risk_status, 0),
            round(total_stake_r, 2),
            round(avg_stake, 2),
            round(avg_odds_r, 2),
            round(avg_ev_r, 2),
            round(profit_r, 2),
            calc_roi(profit_r, total_stake_r),
            pending_r,
            win_r,
            lose_r,
        ])
    style_sheet(ws_risk)

    ws_shifts = wb.create_sheet("Shifts")
    ws_shifts.append(["id", "started_at", "ended_at", "budget", "spent", "status", "bets", "profit", "ROI %"])
    for shift in shifts:
        sid, started_at, ended_at, budget, spent, status = shift
        stats = get_shift_stats(sid)
        count, sum_stake, avg_o, avg_e, w, l, hw, hl, r, p, profit = stats
        ws_shifts.append([sid, started_at, ended_at, budget, spent, status, count, round(profit, 2), calc_roi(profit, sum_stake)])
    style_sheet(ws_shifts)

    ws_bets = wb.create_sheet("Bets")
    ws_bets.append([
        "id", "shift_id", "created_at", "sport", "tournament", "match_name", "match_date",
        "match_time", "match_start_at", "market", "market_group", "odds", "ev",
        "bookmaker", "stake", "result_status", "payout", "profit", "risk_status", "risk_notes"
    ])
    for row in bets:
        row = list(row)
        row[13] = bookmaker_label(row[13])
        ws_bets.append(row)
    style_sheet(ws_bets)

    ws_rej = wb.create_sheet("Risks_Rejected")
    ws_rej.append(["created_at", "match_name", "market", "odds", "bookmaker", "risk_status", "risk_notes", "reason"])
    for row in rejected:
        row = list(row)
        row[4] = bookmaker_label(row[4])
        ws_rej.append(row)
    style_sheet(ws_rej)

    wb.save(filename)
    return filename


def export_shift_to_xlsx(user_id: int, shift_id: int) -> str | None:
    shift = get_shift_by_id(shift_id, user_id)
    if not shift:
        return None

    db = connect()
    bets = db.execute("""
        SELECT id, created_at, sport, tournament, match_name, match_date,
               match_time, match_start_at, market, market_group, odds, ev,
               bookmaker, stake, result_status, payout, profit, risk_status, risk_notes
        FROM bets
        WHERE shift_id = ?
        ORDER BY id ASC
    """, (shift_id,)).fetchall()

    rejected = db.execute("""
        SELECT created_at, match_name, market, odds, bookmaker, risk_status, risk_notes, reason
        FROM rejected_bets
        WHERE user_id = ?
        ORDER BY id DESC
    """, (user_id,)).fetchall()
    db.close()

    if not bets:
        return None

    filename = f"shift_{shift_id}_report_{now_dt().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    sid, uid, started_at, ended_at, budget, spent, status = shift
    total_bets, total_stake, avg_odds, avg_ev, wins, loses, half_wins, half_loses, refunds, pendings, total_profit = get_shift_stats(shift_id)

    ws.append(["Metric", "Value"])
    for row in [
        ["Shift ID", sid],
        ["Started", started_at],
        ["Ended", ended_at or "active"],
        ["Budget", budget],
        ["Spent", spent],
        ["Bets", total_bets],
        ["Total stake", round(total_stake, 2)],
        ["Avg odds", round(avg_odds, 2)],
        ["Avg EV", round(avg_ev, 2)],
        ["Profit", round(total_profit, 2)],
        ["ROI %", calc_roi(total_profit, total_stake)],
        ["Win", wins],
        ["Lose", loses],
        ["Half win", half_wins],
        ["Half lose", half_loses],
        ["Refund", refunds],
        ["Pending", pendings],
    ]:
        ws.append(row)

    ws.append([])
    ws.append(["Market", "Count", "Share %", "Total stake", "Avg stake", "Avg odds", "Avg EV", "Profit", "ROI %", "Pending"])
    market_rows, total = get_market_stats_by_shift(shift_id)
    for market, count, total_stake_m, avg_stake, avg_odds_m, avg_ev_m, profit_m, pending_m in market_rows:
        ws.append([
            market,
            count,
            round(count / total * 100, 2) if total else 0,
            round(total_stake_m, 2),
            round(avg_stake, 2),
            round(avg_odds_m, 2),
            round(avg_ev_m, 2),
            round(profit_m, 2),
            calc_roi(profit_m, total_stake_m),
            pending_m,
        ])

    ws.append([])
    ws.append(["Bookmaker", "Count", "Share %", "Total stake", "Avg stake", "Avg odds", "Profit", "ROI %"])
    book_rows, total_bk = get_bookmaker_stats_by_shift(shift_id)
    for bookmaker, count, total_stake_b, avg_stake, avg_odds_b, profit_b in book_rows:
        ws.append([
            bookmaker_label(bookmaker),
            count,
            round(count / total_bk * 100, 2) if total_bk else 0,
            round(total_stake_b, 2),
            round(avg_stake, 2),
            round(avg_odds_b, 2),
            round(profit_b, 2),
            calc_roi(profit_b, total_stake_b),
        ])

    ws.append([])
    ws.append(["Risk status", "Count", "Share %", "Total stake", "Avg stake", "Avg odds", "Avg EV", "Profit", "ROI %", "Pending", "Win", "Lose"])
    risk_rows, total_risk = get_risk_stats_by_shift(shift_id)
    for risk_status, count, total_stake_r, avg_stake, avg_odds_r, avg_ev_r, profit_r, pending_r, win_r, lose_r in risk_rows:
        ws.append([
            risk_status,
            count,
            round(count / total_risk * 100, 2) if total_risk else 0,
            round(total_stake_r, 2),
            round(avg_stake, 2),
            round(avg_odds_r, 2),
            round(avg_ev_r, 2),
            round(profit_r, 2),
            calc_roi(profit_r, total_stake_r),
            pending_r,
            win_r,
            lose_r,
        ])

    add_summary_charts(ws)
    style_sheet(ws)

    ws_risk = wb.create_sheet("Risk_Analytics")
    ws_risk.append(["Risk status", "Count", "Share %", "Total stake", "Avg stake", "Avg odds", "Avg EV", "Profit", "ROI %", "Pending", "Win", "Lose"])
    for risk_status, count, total_stake_r, avg_stake, avg_odds_r, avg_ev_r, profit_r, pending_r, win_r, lose_r in risk_rows:
        ws_risk.append([
            risk_status,
            count,
            round(count / total_risk * 100, 2) if total_risk else 0,
            round(total_stake_r, 2),
            round(avg_stake, 2),
            round(avg_odds_r, 2),
            round(avg_ev_r, 2),
            round(profit_r, 2),
            calc_roi(profit_r, total_stake_r),
            pending_r,
            win_r,
            lose_r,
        ])
    style_sheet(ws_risk)

    ws_bets = wb.create_sheet("Bets")
    ws_bets.append([
        "id", "created_at", "sport", "tournament", "match_name", "match_date",
        "match_time", "match_start_at", "market", "market_group", "odds", "ev",
        "bookmaker", "stake", "result_status", "payout", "profit", "risk_status", "risk_notes"
    ])
    for bet in bets:
        row = list(bet)
        row[12] = bookmaker_label(row[12])
        ws_bets.append(row)
    style_sheet(ws_bets)

    ws_rejected = wb.create_sheet("Risks_Rejected")
    ws_rejected.append(["created_at", "match_name", "market", "odds", "bookmaker", "risk_status", "risk_notes", "reason"])
    for item in rejected:
        row = list(item)
        row[4] = bookmaker_label(row[4])
        ws_rejected.append(row)
    style_sheet(ws_rejected)

    wb.save(filename)
    return filename


# =========================
# KEYBOARDS
# =========================
def main_menu_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🎯 Смена"), KeyboardButton(text="📚 Ставки")],
            [KeyboardButton(text="📊 Статистика"), KeyboardButton(text="⚙️ Сервис")],
            [KeyboardButton(text="❌ Отмена")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def shift_menu_kb(active: bool):
    keyboard = [[KeyboardButton(text="🚀 Начать смену")]] if not active else [
        [KeyboardButton(text="📍 Текущая смена"), KeyboardButton(text="🏁 Завершить смену")]
    ]
    keyboard.append([KeyboardButton(text="⬅️ Назад")])
    return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True, is_persistent=True)


def bets_menu_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="➕ Добавить ставку")],
            [KeyboardButton(text="📂 Выбрать смену")],
            [KeyboardButton(text="🧾 Последняя ставка")],
            [KeyboardButton(text="⬅️ Назад")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def stats_menu_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📋 Список смен")],
            [KeyboardButton(text="📌 Ближайшие матчи")],
            [KeyboardButton(text="📦 Export XLSX all"), KeyboardButton(text="📤 Export CSV all")],
            [KeyboardButton(text="⬅️ Назад")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def service_menu_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📋 Логи")],
            [KeyboardButton(text="⬅️ Назад")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def yes_no_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="✅ Подтвердить"), KeyboardButton(text="❌ Отмена")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def risk_decision_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=str(FAST_AMOUNTS[0])), KeyboardButton(text=str(FAST_AMOUNTS[1])), KeyboardButton(text=str(FAST_AMOUNTS[2]))],
            [KeyboardButton(text="✍️ Другая сумма"), KeyboardButton(text="🚫 Отказаться")],
            [KeyboardButton(text="❌ Отмена")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def amount_retry_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=str(FAST_AMOUNTS[0])), KeyboardButton(text=str(FAST_AMOUNTS[1])), KeyboardButton(text=str(FAST_AMOUNTS[2]))],
            [KeyboardButton(text="🔁 Повторить ввод суммы")],
            [KeyboardButton(text="❌ Отмена")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def result_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🕒 В ожидании")],
            [KeyboardButton(text="✅ Выигрыш"), KeyboardButton(text="❌ Проигрыш")],
            [KeyboardButton(text="🟡 Половина выигрыша"), KeyboardButton(text="🟠 Половина проигрыша")],
            [KeyboardButton(text="↩️ Возврат")],
            [KeyboardButton(text="❌ Отмена")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def shift_list_kb(has_prev: bool, has_next: bool):
    row = []
    if has_prev:
        row.append(KeyboardButton(text="⬅️ Пред. смены"))
    if has_next:
        row.append(KeyboardButton(text="➡️ След. смены"))

    keyboard = [row] if row else []
    keyboard.append([KeyboardButton(text="⬅️ Назад")])

    return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True, is_persistent=True)


def selected_shift_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📈 Отчет смены"), KeyboardButton(text="📚 Ставки смены")],
            [KeyboardButton(text="📦 XLSX смены")],
            [KeyboardButton(text="⬅️ Назад")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


def selected_bet_action_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="✏️ Изменить сумму выбранной")],
            [KeyboardButton(text="🏷 Рассчитать выбранную")],
            [KeyboardButton(text="⬅️ Назад")],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


# =========================
# FORMATTERS
# =========================
def format_shift_stats_text(shift_id: int, title: str = "📈 <b>Статистика по смене</b>") -> str:
    shift = get_shift_by_id(shift_id, OWNER_ID)
    if not shift:
        return "📭 Смена не найдена."

    sid, uid, started_at, ended_at, budget, spent, status = shift
    total_bets, total_stake, avg_odds, avg_ev, wins, loses, half_wins, half_loses, refunds, pendings, total_profit = get_shift_stats(shift_id)

    remain = round(budget - spent, 2)
    roi = calc_roi(total_profit, total_stake)

    text = (
        f"{title}\n"
        "━━━━━━━━━━━━━━\n"
        f"🆔 Смена: <b>{shift_id}</b>\n"
        f"🕒 Начало: <b>{started_at} МСК</b>\n"
        f"🏁 Конец: <b>{ended_at or 'активна'}</b>\n\n"
        f"🎯 Ставок: <b>{total_bets}</b>\n"
        f"💸 Общая сумма: <b>{round(total_stake, 2)}</b>\n"
        f"📈 Средний КФ: <b>{round(avg_odds, 2) if total_bets else 0}</b>\n"
        f"🧠 Среднее EV: <b>{round(avg_ev, 2) if total_bets else 0}</b>\n\n"
        f"✅ Выигрыш: <b>{wins}</b>\n"
        f"❌ Проигрыш: <b>{loses}</b>\n"
        f"🟡 Half win: <b>{half_wins}</b>\n"
        f"🟠 Half lose: <b>{half_loses}</b>\n"
        f"↩️ Возврат: <b>{refunds}</b>\n"
        f"🕒 Pending: <b>{pendings}</b>\n\n"
        f"💰 Бюджет: <b>{budget}</b>\n"
        f"💸 Поставлено: <b>{spent}</b>\n"
        f"🟢 Остаток: <b>{remain}</b>\n"
        f"📊 Прибыль: <b>{round(total_profit, 2)}</b>\n"
        f"📐 ROI: <b>{roi}%</b>"
    )

    market_rows, total = get_market_stats_by_shift(shift_id)
    if market_rows:
        text += "\n\n📌 <b>Маркеты</b>"
        for market, count, total_stake_m, avg_stake, avg_odds_m, avg_ev_m, profit_m, pending_m in market_rows[:10]:
            share = round(count / total * 100, 2) if total else 0
            text += (
                f"\n• <b>{market}</b>: {count} ставок, {share}%, "
                f"ср. сумма {round(avg_stake, 2)}, ср. КФ {round(avg_odds_m, 2)}, "
                f"ROI {calc_roi(profit_m, total_stake_m)}%"
            )

    return text


def format_bet_card(bet_row) -> str:
    (
        bet_id,
        sport,
        tournament,
        match_name,
        match_date,
        match_time,
        market,
        odds,
        ev,
        bookmaker,
        stake,
        created_at,
        result_status,
        payout,
        profit,
        shift_id,
    ) = bet_row

    return (
        "🎯 <b>Выбрана ставка</b>\n"
        "━━━━━━━━━━━━━━\n"
        f"🆔 ID: <b>{bet_id}</b>\n"
        f"🏟 <b>{match_name}</b>\n\n"
        f"📌 {market}\n\n"
        f"💸 Сумма: <b>{stake}</b>\n"
        f"📈 КФ: <b>{odds}</b>\n"
        f"🏦 {bookmaker_label(bookmaker)}\n"
        f"🏷 Сейчас: <b>{RESULT_LABELS.get(result_status, result_status)}</b>"
    )


def format_fast_scoring(parsed: dict) -> str:
    match_start = datetime.fromisoformat(parsed["match_start_at"]).astimezone(TIMEZONE).strftime("%d.%m.%Y %H:%M")
    risk_text = risk_note_to_telegram(parsed["risk_status"], parsed["risk_notes"])

    return (
        "⚡ <b>БОЕВОЙ СКОРИНГ</b>\n"
        "━━━━━━━━━━━━━━\n"
        f"🏟 <b>{parsed['match_name']}</b>\n"
        f"📌 {parsed['market']}\n"
        f"📊 <b>{parsed.get('market_group')}</b> | 📈 <b>{parsed['odds']}</b> | 🧠 <b>{parsed['ev'] if parsed['ev'] is not None else '-'}</b>\n"
        f"🏦 {bookmaker_label(parsed['bookmaker'])} | 🕒 <b>{match_start} МСК</b>\n\n"
        f"{risk_text}\n\n"
        f"<b>Сумма = подтверждение.</b>\n"
        f"Выбери кнопку или напиши свою сумму."
    )


# =========================
# COMMANDS / NAVIGATION
# =========================
@dp.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID:
        await message.answer("⛔ Этот бот доступен только владельцу.")
        return

    await state.clear()
    await message.answer(
        "🚀 <b>Бот учёта ставок + риск-сканер запущен</b>\n\nВыбери раздел 👇",
        reply_markup=main_menu_kb(),
    )


@dp.message(F.text == "⬅️ Назад")
async def back_to_main(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("🏠 Главное меню", reply_markup=main_menu_kb())


@dp.message(F.text == "❌ Отмена")
async def cancel_action(message: Message, state: FSMContext):
    data = await state.get_data()
    parsed = data.get("pending_bet")
    if parsed:
        save_rejected_bet(message.from_user.id, parsed, "cancelled")

    await state.clear()
    await message.answer("❌ Действие отменено.", reply_markup=main_menu_kb())


@dp.message(F.text == "🎯 Смена")
async def open_shift_menu(message: Message, state: FSMContext):
    await state.clear()
    active = get_active_shift(message.from_user.id)

    if active:
        shift_id, budget, spent, started_at = active
        await message.answer(
            f"🟢 <b>Смена активна</b>\n\n"
            f"🆔 ID: <b>{shift_id}</b>\n"
            f"🕒 {started_at} МСК\n"
            f"💰 Бюджет: <b>{budget}</b>\n"
            f"💸 Поставлено: <b>{spent}</b>\n"
            f"🟢 Остаток: <b>{round(budget - spent, 2)}</b>",
            reply_markup=shift_menu_kb(True),
        )
    else:
        await message.answer("🎯 <b>Раздел «Смена»</b>", reply_markup=shift_menu_kb(False))


@dp.message(F.text == "📚 Ставки")
async def open_bets_menu(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("📚 <b>Раздел «Ставки»</b>", reply_markup=bets_menu_kb())


@dp.message(F.text == "📊 Статистика")
async def open_stats_menu(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("📊 <b>Раздел «Статистика»</b>", reply_markup=stats_menu_kb())


@dp.message(F.text == "⚙️ Сервис")
async def open_service_menu(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("⚙️ <b>Раздел «Сервис»</b>", reply_markup=service_menu_kb())


# =========================
# SHIFT FLOW
# =========================
@dp.message(F.text == "🚀 Начать смену")
async def start_shift_button(message: Message, state: FSMContext):
    if has_recent_action(message.from_user.id, "start_shift"):
        return

    if get_active_shift(message.from_user.id):
        await message.answer("ℹ️ Смена уже активна.", reply_markup=shift_menu_kb(True))
        return

    await state.set_state(ShiftState.waiting_budget)
    await message.answer(
        "💰 <b>Введи бюджет смены</b>\n\nПример: <code>10000</code>",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="❌ Отмена")]],
            resize_keyboard=True,
            is_persistent=True,
        ),
    )


@dp.message(ShiftState.waiting_budget)
async def budget_input(message: Message, state: FSMContext):
    try:
        budget = as_float(message.text.strip())
        if budget <= 0:
            raise ValueError
    except Exception:
        await message.answer("⚠️ Бюджет не распознан. Введи число, например: <code>10000</code>")
        return

    start_shift_db(message.from_user.id, now_str(), budget)
    await state.clear()
    await message.answer(f"✅ <b>Смена начата</b>\n💰 Бюджет: <b>{budget}</b>", reply_markup=shift_menu_kb(True))


@dp.message(F.text == "📍 Текущая смена")
async def current_shift_handler(message: Message):
    active = get_active_shift(message.from_user.id)
    if not active:
        await message.answer("📭 Активной смены нет.", reply_markup=shift_menu_kb(False))
        return

    await message.answer(format_shift_stats_text(active[0], "📊 <b>Текущая смена</b>"), reply_markup=shift_menu_kb(True))


@dp.message(F.text == "🏁 Завершить смену")
async def end_shift_handler(message: Message, state: FSMContext):
    active = get_active_shift(message.from_user.id)
    if not active:
        await message.answer("📭 Активной смены нет.", reply_markup=shift_menu_kb(False))
        return

    shift_id, budget, spent, started_at = active
    await state.set_state(ShiftState.waiting_end_shift_confirm)
    await message.answer(
        f"🏁 <b>Подтвердить завершение смены?</b>\n\n"
        f"💰 Бюджет: <b>{budget}</b>\n"
        f"💸 Поставлено: <b>{spent}</b>\n"
        f"🟢 Остаток: <b>{round(budget - spent, 2)}</b>",
        reply_markup=yes_no_kb(),
    )


@dp.message(ShiftState.waiting_end_shift_confirm, F.text == "✅ Подтвердить")
async def confirm_end_shift(message: Message, state: FSMContext):
    active = get_active_shift(message.from_user.id)
    if not active:
        await state.clear()
        await message.answer("📭 Активной смены уже нет.", reply_markup=shift_menu_kb(False))
        return

    shift_id, budget, spent, started_at = active
    end_shift_db(shift_id, now_str())
    await state.clear()
    await message.answer(
        f"🏁 <b>Смена завершена</b>\n\n{format_shift_stats_text(shift_id, '📊 <b>Итог смены</b>')}",
        reply_markup=shift_menu_kb(False),
    )


# =========================
# BET ADDING / FAST SCORING
# =========================
@dp.message(F.text == "➕ Добавить ставку")
async def add_bet_hint(message: Message, state: FSMContext):
    await state.clear()

    if not get_active_shift(message.from_user.id):
        await message.answer("⚠️ Сначала начни смену.", reply_markup=shift_menu_kb(False))
        return

    await state.set_state(ShiftState.waiting_forwarded_bet)
    await message.answer(
        "📥 <b>Жду пересланную ставку</b>\n\n"
        "Можно просто переслать ставку сюда — я сразу сделаю риск-скоринг.",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="❌ Отмена")]],
            resize_keyboard=True,
            is_persistent=True,
        ),
    )


@dp.message(F.text == "🚫 Отказаться")
async def risk_reject_handler(message: Message, state: FSMContext):
    data = await state.get_data()
    parsed = data.get("pending_bet")

    if parsed:
        save_rejected_bet(message.from_user.id, parsed, "manual_reject")

    await state.clear()
    await message.answer("🚫 Ставка отклонена и сохранена в rejected_bets.", reply_markup=bets_menu_kb())


@dp.message(F.text == "✍️ Другая сумма")
async def other_amount_handler(message: Message, state: FSMContext):
    if await state.get_state() != ShiftState.waiting_risk_decision.state:
        return

    await state.set_state(ShiftState.waiting_bet_amount)
    await message.answer("💬 Напиши сумму ставки или выбери кнопку.", reply_markup=amount_retry_kb())


@dp.message(F.text == "🔁 Повторить ввод суммы")
async def retry_amount_handler(message: Message, state: FSMContext):
    if await state.get_state() != ShiftState.waiting_bet_amount.state:
        await message.answer("ℹ️ Сейчас нет активного ввода суммы.", reply_markup=bets_menu_kb())
        return

    await message.answer("🔁 Напиши сумму заново. Пример: <code>1500</code>", reply_markup=amount_retry_kb())


async def save_pending_bet_amount(message: Message, state: FSMContext, amount: float):
    data = await state.get_data()
    pending = data.get("pending_bet")

    if not pending:
        await state.clear()
        await message.answer("⚠️ Не нашёл ожидаемую ставку.", reply_markup=bets_menu_kb())
        return

    active = get_active_shift(message.from_user.id)
    if not active:
        await state.clear()
        await message.answer("📭 Активной смены нет.", reply_markup=shift_menu_kb(False))
        return

    shift_id, budget, spent, started_at = active

    try:
        add_bet_db(shift_id, message.from_user.id, now_str(), pending, amount)
    except Exception as e:
        if "UNIQUE constraint failed" in str(e):
            await state.clear()
            await message.answer(
                "⚠️ Эта ставка уже была добавлена ранее.\n\n"
                "Если это новая ставка по тому же событию — проверь, не переслал ли ты тот же самый исходник.",
                reply_markup=bets_menu_kb(),
            )
            return

        log_error(f"Bet insert failed: {e}")
        await state.clear()
        await message.answer(f"❌ Ошибка записи ставки: <code>{e}</code>", reply_markup=bets_menu_kb())
        return

    new_spent = round(spent + amount, 2)
    remain = round(budget - new_spent, 2)
    warn = ""
    if new_spent > budget:
        warn = f"\n\n⚠️ <b>Выход за лимит</b> на <b>{round(new_spent - budget, 2)}</b>"

    await state.clear()
    await message.answer(
        f"✅ <b>Ставка сохранена</b>\n\n"
        f"💸 Сумма: <b>{amount}</b>\n"
        f"📊 Поставлено: <b>{new_spent}</b> / <b>{budget}</b>\n"
        f"🟢 Остаток: <b>{remain}</b>{warn}",
        reply_markup=bets_menu_kb(),
    )


async def show_scored_bet(message: Message, state: FSMContext, parsed: dict):
    parsed = analyze_risk(parsed)

    await state.update_data(pending_bet=parsed)
    await state.set_state(ShiftState.waiting_risk_decision)

    await message.answer(
        format_fast_scoring(parsed),
        reply_markup=risk_decision_kb(),
    )


# =========================
# BETS VIEW
# =========================
@dp.message(F.text == "🧾 Последняя ставка")
async def last_bet_handler(message: Message):
    row = get_last_bet(message.from_user.id)
    if not row:
        await message.answer("📭 Пока нет ставок.", reply_markup=bets_menu_kb())
        return

    bet_id, sport, match_name, market, odds, stake, bookmaker, created_at, result_status = row
    await message.answer(
        "🧾 <b>Последняя ставка</b>\n"
        "━━━━━━━━━━━━━━\n"
        f"🏅 {sport}\n"
        f"🏟 <b>{match_name}</b>\n\n"
        f"📌 {market}\n\n"
        f"📈 КФ: <b>{odds}</b>\n"
        f"💸 Сумма: <b>{stake}</b>\n"
        f"🏦 {bookmaker_label(bookmaker)}\n"
        f"🏷 {RESULT_LABELS.get(result_status, result_status)}",
        reply_markup=bets_menu_kb(),
    )


# =========================
# SHIFT LIST / SELECTED SHIFT
# =========================
@dp.message(F.text.in_({"📋 Список смен", "📂 Выбрать смену"}))
async def shift_list_start(message: Message, state: FSMContext):
    await show_shift_page(message, state, 0)


async def show_shift_page(message: Message, state: FSMContext, page: int):
    offset = page * SHIFT_PAGE_SIZE
    rows, total = list_shifts(message.from_user.id, offset, SHIFT_PAGE_SIZE)

    if not rows:
        await message.answer("📭 Смен пока нет.", reply_markup=stats_menu_kb())
        return

    mapping = {}
    lines = [f"📋 <b>Смены {offset + 1}–{offset + len(rows)} из {total}</b>\n"]

    for idx, row in enumerate(rows, 1):
        shift_id, started_at, ended_at, budget, spent, status = row
        stats = get_shift_stats(shift_id)
        total_bets, total_stake, avg_odds, avg_ev, wins, loses, half_wins, half_loses, refunds, pendings, total_profit = stats

        mapping[str(idx)] = shift_id
        lines.append(
            f"{idx}. ID <b>{shift_id}</b> | {started_at}\n"
            f"Статус: <b>{status}</b>\n"
            f"Бюджет: <b>{budget}</b> | Поставлено: <b>{spent}</b>\n"
            f"Ставок: <b>{total_bets}</b> | ROI: <b>{calc_roi(total_profit, total_stake)}%</b>\n"
        )

    await state.update_data(shift_page=page, shift_choices=mapping)
    await state.set_state(ShiftState.waiting_shift_number)
    await message.answer(
        "\n".join(lines) + "\nНапиши номер смены.",
        reply_markup=shift_list_kb(page > 0, offset + len(rows) < total),
    )


@dp.message(ShiftState.waiting_shift_number, F.text.in_({"➡️ След. смены", "⬅️ Пред. смены"}))
async def shift_page_nav(message: Message, state: FSMContext):
    data = await state.get_data()
    page = data.get("shift_page", 0)

    page = page + 1 if message.text == "➡️ След. смены" else max(0, page - 1)
    await show_shift_page(message, state, page)


@dp.message(ShiftState.waiting_shift_number)
async def shift_choose_number(message: Message, state: FSMContext):
    data = await state.get_data()
    mapping = data.get("shift_choices", {})
    text = (message.text or "").strip()

    if text not in mapping:
        await message.answer("⚠️ Номер смены не найден.")
        return

    shift_id = mapping[text]
    await state.update_data(selected_shift_id=shift_id)
    await state.set_state(None)

    await message.answer(
        format_shift_stats_text(shift_id, f"📊 <b>Выбрана смена {shift_id}</b>"),
        reply_markup=selected_shift_kb(),
    )


@dp.message(F.text == "📈 Отчет смены")
async def selected_shift_report(message: Message, state: FSMContext):
    data = await state.get_data()
    shift_id = data.get("selected_shift_id")

    if not shift_id:
        await message.answer("⚠️ Сначала выбери смену через 📋 Список смен.", reply_markup=stats_menu_kb())
        return

    await message.answer(format_shift_stats_text(shift_id), reply_markup=selected_shift_kb())


@dp.message(F.text == "📚 Ставки смены")
async def selected_shift_bets(message: Message, state: FSMContext):
    data = await state.get_data()
    shift_id = data.get("selected_shift_id")

    if not shift_id:
        await message.answer("⚠️ Сначала выбери смену через 📋 Список смен.", reply_markup=stats_menu_kb())
        return

    rows, total = get_bets_by_shift(shift_id, 0, BET_PAGE_SIZE)
    if not rows:
        await message.answer("📭 В этой смене нет ставок.", reply_markup=selected_shift_kb())
        return

    mapping = {}
    lines = [f"📚 <b>Ставки смены {shift_id}</b>\nПоказано 1–{len(rows)} из {total}\n"]

    for idx, row in enumerate(rows, 1):
        bet_id, sport, match_name, market, odds, stake, bookmaker, created_at, result_status = row
        mapping[str(idx)] = bet_id
        lines.append(
            f"{idx}. <b>{match_name}</b>\n"
            f"📌 {market}\n"
            f"💸 {stake} | 📈 {odds} | 🏦 {bookmaker_label(bookmaker)}\n"
            f"🏷 {RESULT_LABELS.get(result_status, result_status)}\n"
        )

    await state.update_data(shift_bet_choices=mapping, selected_shift_id=shift_id)
    await state.set_state(ShiftState.waiting_shift_bet_number)
    await message.answer(
        "\n".join(lines) + "\nНапиши номер ставки.",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="❌ Отмена")]],
            resize_keyboard=True,
            is_persistent=True,
        ),
    )


@dp.message(ShiftState.waiting_shift_bet_number)
async def selected_shift_bet_number(message: Message, state: FSMContext):
    text = (message.text or "").strip()
    data = await state.get_data()
    mapping = data.get("shift_bet_choices", {})

    if text not in mapping:
        await message.answer("⚠️ Номер ставки не найден. Напиши номер из списка.")
        return

    bet_id = mapping[text]
    bet = get_bet_by_id(bet_id)

    if not bet:
        await state.clear()
        await message.answer("⚠️ Не удалось найти ставку.", reply_markup=selected_shift_kb())
        return

    shift_id = bet[-1]
    await state.update_data(selected_shift_bet_id=bet_id, selected_shift_id=shift_id)
    await state.set_state(ShiftState.waiting_selected_bet_action)

    await message.answer(format_bet_card(bet), reply_markup=selected_bet_action_kb())


@dp.message(ShiftState.waiting_selected_bet_action, F.text == "✏️ Изменить сумму выбранной")
async def selected_bet_edit_amount_start(message: Message, state: FSMContext):
    data = await state.get_data()
    bet_id = data.get("selected_shift_bet_id")

    if not bet_id:
        await state.clear()
        await message.answer("⚠️ Ставка не выбрана.", reply_markup=main_menu_kb())
        return

    await state.set_state(ShiftState.waiting_selected_bet_new_stake)
    await message.answer(
        "✏️ Напиши новую сумму ставки.",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="❌ Отмена")]],
            resize_keyboard=True,
            is_persistent=True,
        ),
    )


@dp.message(ShiftState.waiting_selected_bet_new_stake)
async def selected_bet_edit_amount_finish(message: Message, state: FSMContext):
    data = await state.get_data()
    bet_id = data.get("selected_shift_bet_id")
    shift_id = data.get("selected_shift_id")

    try:
        new_stake = as_float(message.text.strip())
        if new_stake <= 0:
            raise ValueError
    except Exception:
        await message.answer("⚠️ Сумма не распознана. Введи число, например: <code>1500</code>")
        return

    ok, old_stake = update_bet_stake(bet_id, new_stake)
    await state.update_data(selected_shift_id=shift_id)
    await state.set_state(None)

    await message.answer(
        f"✅ <b>Сумма обновлена</b>\n\nСтарая: <b>{old_stake}</b>\nНовая: <b>{new_stake}</b>"
        if ok else f"⚠️ {old_stake}",
        reply_markup=selected_shift_kb(),
    )


@dp.message(ShiftState.waiting_selected_bet_action, F.text == "🏷 Рассчитать выбранную")
async def selected_bet_result_start(message: Message, state: FSMContext):
    data = await state.get_data()
    bet_id = data.get("selected_shift_bet_id")

    if not bet_id:
        await state.clear()
        await message.answer("⚠️ Ставка не выбрана.", reply_markup=main_menu_kb())
        return

    await state.set_state(ShiftState.waiting_selected_bet_result_status)
    await message.answer("🏷 Выбери результат:", reply_markup=result_kb())


@dp.message(ShiftState.waiting_selected_bet_result_status, F.text.in_(list(RESULT_MAP.keys())))
async def selected_bet_result_finish(message: Message, state: FSMContext):
    data = await state.get_data()
    bet_id = data.get("selected_shift_bet_id")
    shift_id = data.get("selected_shift_id")
    result_status = RESULT_MAP[message.text]

    if not bet_id:
        await state.clear()
        await message.answer("⚠️ Ставка не выбрана.", reply_markup=main_menu_kb())
        return

    ok = update_bet_result(bet_id, result_status)
    await state.update_data(selected_shift_id=shift_id)
    await state.set_state(None)

    await message.answer(
        f"✅ Результат обновлён: <b>{RESULT_LABELS[result_status]}</b>"
        if ok else "⚠️ Не удалось обновить результат.",
        reply_markup=selected_shift_kb(),
    )


# =========================
# STATS / EXPORTS
# =========================
@dp.message(F.text == "📤 Export CSV all")
async def export_csv_all(message: Message):
    path = export_bets_to_csv(message.from_user.id)

    if not path:
        await message.answer("📭 Нет данных.", reply_markup=stats_menu_kb())
        return

    await message.answer_document(FSInputFile(path), caption="📤 CSV за всё время готов.", reply_markup=stats_menu_kb())


@dp.message(F.text == "📦 Export XLSX all")
async def export_xlsx_all(message: Message):
    path = export_all_to_xlsx(message.from_user.id)

    if not path:
        await message.answer("📭 Нет данных.", reply_markup=stats_menu_kb())
        return

    await message.answer_document(FSInputFile(path), caption="📦 Общий XLSX отчёт готов.", reply_markup=stats_menu_kb())


@dp.message(F.text == "📦 XLSX смены")
async def selected_shift_xlsx(message: Message, state: FSMContext):
    data = await state.get_data()
    shift_id = data.get("selected_shift_id")

    if not shift_id:
        await message.answer("⚠️ Сначала выбери смену через 📋 Список смен.", reply_markup=stats_menu_kb())
        return

    path = export_shift_to_xlsx(message.from_user.id, shift_id)

    if not path:
        await message.answer("📭 Нет данных для экспорта.", reply_markup=selected_shift_kb())
        return

    await message.answer_document(
        FSInputFile(path),
        caption=f"📦 XLSX отчёт по смене {shift_id} готов.",
        reply_markup=selected_shift_kb(),
    )


@dp.message(F.text == "📌 Ближайшие матчи")
async def upcoming_matches_handler(message: Message):
    db = connect()
    rows = db.execute("""
        SELECT match_name, market, stake, odds, bookmaker, match_start_at
        FROM bets
        WHERE user_id = ?
          AND result_status = 'pending'
          AND match_start_at IS NOT NULL
        ORDER BY match_start_at ASC
        LIMIT 20
    """, (message.from_user.id,)).fetchall()
    db.close()

    future = []
    for match_name, market, stake, odds, bookmaker, match_start_at in rows:
        try:
            dt = datetime.fromisoformat(match_start_at)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=TIMEZONE)
        except Exception:
            continue

        if dt >= now_dt():
            future.append((match_name, market, stake, odds, bookmaker, dt))

    if not future:
        await message.answer("📭 Ближайших pending матчей нет.", reply_markup=stats_menu_kb())
        return

    lines = ["📌 <b>Ближайшие матчи</b>\n"]
    for i, (match_name, market, stake, odds, bookmaker, dt) in enumerate(future[:10], 1):
        lines.append(
            f"{i}. <b>{match_name}</b>\n"
            f"📌 {market}\n"
            f"💸 {stake} | 📈 {odds} | 🏦 {bookmaker_label(bookmaker)}\n"
            f"🕒 {dt.strftime('%d.%m.%Y %H:%M')} МСК\n"
        )

    await message.answer("\n".join(lines), reply_markup=stats_menu_kb())


# =========================
# SERVICE
# =========================
@dp.message(F.text == "📋 Логи")
async def logs_handler(message: Message):
    rows = get_recent_logs(10)

    if not rows:
        await message.answer("📭 Логов нет.", reply_markup=service_menu_kb())
        return

    lines = ["📋 <b>Последние логи</b>\n"]
    for level, text, created_at in rows:
        lines.append(f"<b>{level}</b> | {created_at}\n{text}\n")

    await message.answer("\n".join(lines), reply_markup=service_menu_kb())


# =========================
# UNIVERSAL TEXT / DIRECT BET FLOW
# =========================
@dp.message()
async def universal_message_handler(message: Message, state: FSMContext):
    """
    Главный catch-all обработчик:
    - прямой пересыл ставки работает из любого обычного меню;
    - сумма после скоринга сразу сохраняет ставку;
    - если вместо суммы пришла новая ставка, старая pending-заявка переносится в rejected.
    """
    if not message.from_user or message.from_user.id != OWNER_ID:
        return

    text = (message.text or message.caption or "").strip()
    if not text:
        return

    current_state = await state.get_state()

    parsed = parse_bet(text)

    if parsed:
        active = get_active_shift(message.from_user.id)
        if not active:
            await message.answer("⚠️ Сначала начни смену.", reply_markup=shift_menu_kb(False))
            return

        if current_state in {ShiftState.waiting_risk_decision.state, ShiftState.waiting_bet_amount.state}:
            old_data = await state.get_data()
            old_pending = old_data.get("pending_bet")
            if old_pending:
                save_rejected_bet(message.from_user.id, old_pending, "replaced_by_new_forward")

        await show_scored_bet(message, state, parsed)
        return

    if current_state == ShiftState.waiting_forwarded_bet.state:
        diagnostics = parse_bet_diagnostics(text)
        await message.answer(
            "⚠️ <b>Ставка не распознана</b>\n\n"
            "<b>Что не получилось определить:</b>\n"
            f"{diagnostics}\n\n"
            "<b>Что сделать:</b>\n"
            "Перешли исходное сообщение без ручного редактирования.",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="❌ Отмена")]],
                resize_keyboard=True,
                is_persistent=True,
            ),
        )
        log_warning(f"Bet parse failed | diagnostics: {diagnostics}")
        await state.set_state(ShiftState.waiting_forwarded_bet)
        return

    if current_state == ShiftState.waiting_risk_decision.state:
        try:
            amount = as_float(text)
            if amount <= 0:
                raise ValueError
        except Exception:
            await message.answer(
                f"ℹ️ Выбери сумму кнопкой <b>{FAST_AMOUNTS[0]} / {FAST_AMOUNTS[1]} / {FAST_AMOUNTS[2]}</b>, "
                f"напиши свою сумму или нажми <b>🚫 Отказаться</b>.",
                reply_markup=risk_decision_kb(),
            )
            return

        await save_pending_bet_amount(message, state, amount)
        return

    if current_state == ShiftState.waiting_bet_amount.state:
        try:
            amount = as_float(text)
            if amount <= 0:
                raise ValueError
        except Exception:
            await message.answer(
                "⚠️ Сумма не распознана. Введи число, например: <code>1500</code>",
                reply_markup=amount_retry_kb(),
            )
            return

        await save_pending_bet_amount(message, state, amount)
        return

    locked_states = {
        ShiftState.waiting_budget.state,
        ShiftState.waiting_end_shift_confirm.state,
        ShiftState.waiting_shift_number.state,
        ShiftState.waiting_shift_bet_number.state,
        ShiftState.waiting_selected_bet_action.state,
        ShiftState.waiting_selected_bet_new_stake.state,
        ShiftState.waiting_selected_bet_result_status.state,
    }

    if current_state in locked_states:
        return


# =========================
# REMINDERS
# =========================
async def reminder_job():
    reminders = get_due_reminders()

    for item in reminders:
        dt_text = item["match_start_at"].astimezone(TIMEZONE).strftime("%d.%m.%Y %H:%M")
        try:
            await bot.send_message(
                item["user_id"],
                "⏰ <b>Напоминание</b>\n\n"
                f"Через {REMINDER_MINUTES} минут матч:\n"
                f"🏟 <b>{item['match_name']}</b>\n"
                f"📌 {item['market']}\n"
                f"💸 {item['stake']}\n"
                f"📈 {item['odds']}\n"
                f"🏦 {bookmaker_label(item['bookmaker'])}\n"
                f"🕒 {dt_text} МСК",
                reply_markup=bets_menu_kb(),
            )
            mark_reminder_sent(item["id"])
            log_info(f"Reminder sent | bet_id={item['id']}")
        except Exception as e:
            log_error(f"Reminder failed | bet_id={item['id']} | error={e}")


# =========================
# STARTUP
# =========================
async def main():
    print("BOT STARTED")
    init_db()
    scheduler.add_job(reminder_job, "interval", seconds=30, max_instances=1, coalesce=True)
    scheduler.start()
    log_info("Bot started")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
