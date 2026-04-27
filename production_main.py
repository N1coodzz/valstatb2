
import asyncio
import csv
import hashlib
import logging
import os
import re
import sqlite3
import time
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import FSInputFile, KeyboardButton, Message, ReplyKeyboardMarkup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# =========================
# CONFIG
# =========================
BOT_TOKEN = os.getenv("BOT_TOKEN")
OWNER_ID = int(os.getenv("OWNER_ID", "0"))

TIMEZONE = ZoneInfo("Europe/Moscow")
DB_PATH = "bot.db"
LOG_PATH = "bot.log"

REMINDER_MINUTES = 10
DOUBLE_TAP_SECONDS = 1.2
SHIFT_PAGE_SIZE = 5
BET_PAGE_SIZE = 15

# Источник присылает время матча на 1 час позже МСК.
SOURCE_TIME_AHEAD_HOURS = 1

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN not found in environment")
if not OWNER_ID:
    raise RuntimeError("OWNER_ID not found in environment")


# =========================
# LOGGING / BOT
# =========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.FileHandler(LOG_PATH, encoding="utf-8"), logging.StreamHandler()],
)
logger = logging.getLogger("value-bot")

bot = Bot(BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())
scheduler = AsyncIOScheduler(timezone="Europe/Moscow")
ACTION_GUARD = {}


# =========================
# STATES
# =========================
class S(StatesGroup):
    waiting_budget = State()
    waiting_risk_decision = State()
    waiting_bet_amount = State()
    waiting_end_shift_confirm = State()
    waiting_shift_number = State()
    waiting_shift_bet_number = State()
    waiting_selected_bet_action = State()
    waiting_selected_bet_new_stake = State()
    waiting_selected_bet_result_status = State()


# =========================
# CONSTANTS
# =========================
RESULT_MAP = {
    "🕒 В ожидании": "pending",
    "✅ Выигрыш": "win",
    "❌ Проигрыш": "lose",
    "🟡 Половина выигрыша": "half_win",
    "🟠 Половина проигрыша": "half_lose",
    "↩️ Возврат": "refund",
}
RESULT_LABELS = {
    "pending": "🕒 В ожидании",
    "win": "✅ Выигрыш",
    "lose": "❌ Проигрыш",
    "half_win": "🟡 Половина выигрыша",
    "half_lose": "🟠 Половина проигрыша",
    "refund": "↩️ Возврат",
}
BOOKMAKER_EMOJI = {
    "fonbet": "🔴", "фонбет": "🔴",
    "betcity": "🔵", "бетсити": "🔵",
    "betboom": "🟣", "бетбум": "🟣",
    "marathon": "🟠", "марафон": "🟠",
    "ligastavok": "🟢", "лига ставок": "🟢", "liga stavok": "🟢",
}


# =========================
# HELPERS
# =========================
def now_dt() -> datetime:
    return datetime.now(TIMEZONE)


def now_str() -> str:
    return now_dt().strftime("%Y-%m-%d %H:%M:%S")


def as_float(text: str) -> float:
    return float((text or "").replace(" ", "").replace(",", "."))


def db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def one(sql: str, args=()):
    conn = db()
    row = conn.execute(sql, args).fetchone()
    conn.close()
    return row


def all_rows(sql: str, args=()):
    conn = db()
    rows = conn.execute(sql, args).fetchall()
    conn.close()
    return rows


def exec_sql(sql: str, args=()):
    conn = db()
    conn.execute(sql, args)
    conn.commit()
    conn.close()


def has_recent_action(user_id: int, action: str, seconds: float = DOUBLE_TAP_SECONDS) -> bool:
    key = f"{user_id}:{action}"
    now_ts = time.time()
    last = ACTION_GUARD.get(key)
    ACTION_GUARD[key] = now_ts
    return last is not None and (now_ts - last) < seconds


def is_forward_message(message: Message) -> bool:
    return bool(
        getattr(message, "forward_origin", None)
        or getattr(message, "forward_from_chat", None)
        or getattr(message, "forward_from", None)
        or getattr(message, "forward_sender_name", None)
    )


def normalize_text(text: str) -> str:
    text = (text or "").lower().replace("ё", "е").replace("−", "-").replace("–", "-")
    text = re.sub(r"[^a-zа-я0-9+\-. ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def hash_text(text: str) -> str:
    return hashlib.md5((text or "").strip().encode("utf-8")).hexdigest()


def bookmaker_label(bookmaker: str) -> str:
    n = normalize_text(bookmaker)
    for key, emoji in BOOKMAKER_EMOJI.items():
        if key in n:
            return f"{emoji} {bookmaker}"
    return f"⚪ {bookmaker or 'Other'}"


def split_sport_tournament(header: str):
    parts = [x.strip() for x in re.split(r"\s+-\s+", header.strip()) if x.strip()]
    return (parts[0] if parts else ""), (" - ".join(parts[1:]) if len(parts) > 1 else "")


def split_teams(match_name: str):
    parts = re.split(r"\s+[–-]\s+", match_name or "")
    return (parts[0].strip(), parts[1].strip()) if len(parts) >= 2 else ("", "")


def token_overlap_score(phrase: str, team: str) -> int:
    a = {t for t in normalize_text(phrase).split() if len(t) >= 3}
    b = {t for t in normalize_text(team).split() if len(t) >= 3}
    return len(a & b)


def infer_selection_side(selection: str, team_a: str, team_b: str) -> str:
    s = normalize_text(selection)
    if re.search(r"\bп1\b|\b1\b", s):
        return "home"
    if re.search(r"\bп2\b|\b2\b", s):
        return "away"
    a = token_overlap_score(selection, team_a)
    b = token_overlap_score(selection, team_b)
    if a > b and a > 0:
        return "home"
    if b > a and b > 0:
        return "away"
    return f"selection:{normalize_text(selection)[:60]}"


def extract_selection_phrase(market: str) -> str:
    m = re.search(r"(.+?)\s+(?:победит\s+с\s+форой|с\s+форой|фора|гандикап)", market, re.I)
    if m:
        return m.group(1).strip(" .,-")
    m = re.search(r"победа\s+(.+?)(?:\s+с\s+учетом|\s+-|\.|$)", market, re.I)
    if m:
        return m.group(1).strip(" .,-")
    return (market or "")[:80]


def parse_line_value(text: str):
    text = (text or "").replace("−", "-").replace(",", ".")
    m = re.search(r"(?:фор[ао]й?|гандикап)[^+\-0-9]{0,20}([+\-]?\d+(?:\.\d+)?)", text, re.I)
    if m:
        return float(m.group(1))
    m = re.search(r"тотал\s+(?:больше|меньше)\s+(\d+(?:\.\d+)?)", text, re.I)
    if m:
        return float(m.group(1))
    return None


def parse_match_start(match_date: str, match_time: str) -> datetime:
    day, month = match_date.split("/")
    hour, minute = match_time.split(":")
    now = now_dt()
    dt = datetime(now.year, int(month), int(day), int(hour), int(minute), tzinfo=TIMEZONE)
    dt = dt - timedelta(hours=SOURCE_TIME_AHEAD_HOURS)
    if dt < now - timedelta(days=30):
        dt = dt.replace(year=now.year + 1)
    return dt


def event_key_from(match_name: str, match_start_at: str) -> str:
    return f"{normalize_text(match_name)}__{match_start_at[:16]}"


def calc_settlement(stake: float, odds: float, result_status: str):
    if result_status == "pending":
        return None, None
    if result_status == "win":
        payout = round(stake * odds, 2)
        return payout, round(payout - stake, 2)
    if result_status == "lose":
        return 0.0, round(-stake, 2)
    if result_status == "half_win":
        payout = round((stake / 2) * odds + (stake / 2), 2)
        return payout, round(payout - stake, 2)
    if result_status == "half_lose":
        payout = round(stake / 2, 2)
        return payout, round(payout - stake, 2)
    if result_status == "refund":
        return round(stake, 2), 0.0
    return None, None


def calc_roi(profit: float, stake: float) -> float:
    return 0.0 if not stake else round((profit / stake) * 100, 2)


def hedge_amount(stake1: float, odds1: float, odds2: float) -> float:
    return 0.0 if not odds2 else round((stake1 * odds1) / odds2, 2)


def arbitrage_metrics(stake1: float, odds1: float, odds2: float):
    stake2 = hedge_amount(stake1, odds1, odds2)
    payout = round(stake1 * odds1, 2)
    total_stake = round(stake1 + stake2, 2)
    profit = round(payout - total_stake, 2)
    roi = calc_roi(profit, total_stake)
    implied = round(1 / odds1 + 1 / odds2, 4) if odds1 and odds2 else 0
    return stake2, payout, total_stake, profit, roi, implied


def handicap_bound(side: str, line: float):
    if side == "home":
        return "gt", -line
    if side == "away":
        return "lt", line
    return None, None


def save_log(level: str, message: str):
    conn = db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS logs(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            level TEXT NOT NULL,
            message TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    conn.execute("INSERT INTO logs(level, message, created_at) VALUES (?, ?, ?)", (level, message, now_str()))
    conn.commit()
    conn.close()


def log_info(text: str):
    logger.info(text)
    save_log("INFO", text)


def log_warning(text: str):
    logger.warning(text)
    save_log("WARNING", text)


def log_error(text: str):
    logger.error(text)
    save_log("ERROR", text)


# =========================
# DATABASE
# =========================
def add_column_if_not_exists(table: str, column: str, ddl: str):
    conn = db()
    cols = [c["name"] for c in conn.execute(f"PRAGMA table_info({table})").fetchall()]
    if column not in cols:
        conn.execute(ddl)
        conn.commit()
    conn.close()


def init_db():
    conn = db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS shifts(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            started_at TEXT NOT NULL,
            ended_at TEXT,
            budget REAL NOT NULL,
            spent REAL NOT NULL DEFAULT 0,
            status TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS bets(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shift_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            sport TEXT,
            tournament TEXT,
            match_name TEXT,
            match_date TEXT,
            match_time TEXT,
            match_start_at TEXT,
            market TEXT,
            odds REAL,
            ev REAL,
            bookmaker TEXT,
            stake REAL,
            source_text TEXT,
            match_hash TEXT UNIQUE,
            reminder_sent INTEGER DEFAULT 0,
            result_status TEXT DEFAULT 'pending',
            payout REAL,
            profit REAL,
            event_key TEXT,
            team_a TEXT,
            team_b TEXT,
            selection_name TEXT,
            selection_side TEXT,
            market_type TEXT,
            market_group TEXT,
            market_side TEXT,
            line_value REAL,
            period_type TEXT,
            semantic_key TEXT,
            risk_status TEXT DEFAULT 'new',
            risk_notes TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS rejected_bets(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            sport TEXT,
            tournament TEXT,
            match_name TEXT,
            market TEXT,
            odds REAL,
            ev REAL,
            bookmaker TEXT,
            source_text TEXT,
            event_key TEXT,
            market_type TEXT,
            market_group TEXT,
            risk_status TEXT,
            risk_notes TEXT,
            reason TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS logs(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            level TEXT NOT NULL,
            message TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

    for table, col, ddl in [
        ("bets", "match_start_at", "ALTER TABLE bets ADD COLUMN match_start_at TEXT"),
        ("bets", "reminder_sent", "ALTER TABLE bets ADD COLUMN reminder_sent INTEGER DEFAULT 0"),
        ("bets", "result_status", "ALTER TABLE bets ADD COLUMN result_status TEXT DEFAULT 'pending'"),
        ("bets", "payout", "ALTER TABLE bets ADD COLUMN payout REAL"),
        ("bets", "profit", "ALTER TABLE bets ADD COLUMN profit REAL"),
        ("bets", "event_key", "ALTER TABLE bets ADD COLUMN event_key TEXT"),
        ("bets", "team_a", "ALTER TABLE bets ADD COLUMN team_a TEXT"),
        ("bets", "team_b", "ALTER TABLE bets ADD COLUMN team_b TEXT"),
        ("bets", "selection_name", "ALTER TABLE bets ADD COLUMN selection_name TEXT"),
        ("bets", "selection_side", "ALTER TABLE bets ADD COLUMN selection_side TEXT"),
        ("bets", "market_type", "ALTER TABLE bets ADD COLUMN market_type TEXT"),
        ("bets", "market_group", "ALTER TABLE bets ADD COLUMN market_group TEXT"),
        ("bets", "market_side", "ALTER TABLE bets ADD COLUMN market_side TEXT"),
        ("bets", "line_value", "ALTER TABLE bets ADD COLUMN line_value REAL"),
        ("bets", "period_type", "ALTER TABLE bets ADD COLUMN period_type TEXT"),
        ("bets", "semantic_key", "ALTER TABLE bets ADD COLUMN semantic_key TEXT"),
        ("bets", "risk_status", "ALTER TABLE bets ADD COLUMN risk_status TEXT DEFAULT 'new'"),
        ("bets", "risk_notes", "ALTER TABLE bets ADD COLUMN risk_notes TEXT"),
    ]:
        add_column_if_not_exists(table, col, ddl)
    log_info("Database initialized")


def get_active_shift(user_id: int):
    return one("SELECT id, budget, spent, started_at FROM shifts WHERE user_id=? AND status='active' ORDER BY id DESC LIMIT 1", (user_id,))


def start_shift_db(user_id: int, budget: float):
    exec_sql("INSERT INTO shifts(user_id, started_at, budget, spent, status) VALUES (?, ?, ?, 0, 'active')", (user_id, now_str(), budget))
    log_info(f"Shift started | user={user_id} | budget={budget}")


def end_shift_db(shift_id: int):
    exec_sql("UPDATE shifts SET ended_at=?, status='ended' WHERE id=?", (now_str(), shift_id))
    log_info(f"Shift ended | shift_id={shift_id}")


def get_shift_by_id(shift_id: int, user_id: int):
    return one("SELECT id, user_id, started_at, ended_at, budget, spent, status FROM shifts WHERE id=? AND user_id=?", (shift_id, user_id))


def list_shifts(user_id: int, offset: int = 0, limit: int = SHIFT_PAGE_SIZE):
    rows = all_rows("SELECT id, started_at, ended_at, budget, spent, status FROM shifts WHERE user_id=? ORDER BY id DESC LIMIT ? OFFSET ?", (user_id, limit, offset))
    total = one("SELECT COUNT(*) AS c FROM shifts WHERE user_id=?", (user_id,))["c"]
    return rows, total


def add_bet_db(shift_id: int, user_id: int, parsed: dict, stake: float):
    conn = db()
    conn.execute("""
        INSERT INTO bets(
            shift_id, user_id, created_at, sport, tournament, match_name, match_date, match_time, match_start_at,
            market, odds, ev, bookmaker, stake, source_text, match_hash,
            reminder_sent, result_status, payout, profit,
            event_key, team_a, team_b, selection_name, selection_side,
            market_type, market_group, market_side, line_value, period_type,
            semantic_key, risk_status, risk_notes
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 'pending', NULL, NULL,
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        shift_id, user_id, now_str(), parsed["sport"], parsed["tournament"], parsed["match_name"], parsed["match_date"],
        parsed["match_time"], parsed["match_start_at"], parsed["market"], parsed["odds"], parsed["ev"], parsed["bookmaker"],
        stake, parsed["source_text"], parsed["hash"], parsed.get("event_key"), parsed.get("team_a"), parsed.get("team_b"),
        parsed.get("selection_name"), parsed.get("selection_side"), parsed.get("market_type"), parsed.get("market_group"),
        parsed.get("market_side"), parsed.get("line_value"), parsed.get("period_type"), parsed.get("semantic_key"),
        parsed.get("risk_status", "new"), parsed.get("risk_notes", ""),
    ))
    conn.execute("UPDATE shifts SET spent = spent + ? WHERE id = ?", (stake, shift_id))
    conn.commit()
    conn.close()
    log_info(f"Bet added | shift_id={shift_id} | user={user_id} | stake={stake} | risk={parsed.get('risk_status')}")


def save_rejected_bet(user_id: int, parsed: dict, reason: str):
    conn = db()
    conn.execute("""
        INSERT INTO rejected_bets(created_at, user_id, sport, tournament, match_name, market, odds, ev, bookmaker,
                                  source_text, event_key, market_type, market_group, risk_status, risk_notes, reason)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        now_str(), user_id, parsed.get("sport"), parsed.get("tournament"), parsed.get("match_name"), parsed.get("market"),
        parsed.get("odds"), parsed.get("ev"), parsed.get("bookmaker"), parsed.get("source_text"), parsed.get("event_key"),
        parsed.get("market_type"), parsed.get("market_group"), parsed.get("risk_status"), parsed.get("risk_notes"), reason
    ))
    conn.commit()
    conn.close()
    log_warning(f"Rejected bet saved | reason={reason} | risk={parsed.get('risk_status')}")


def get_pending_similar_bets(parsed: dict):
    return all_rows("""
        SELECT id, sport, match_name, market, odds, stake, bookmaker, result_status,
               selection_side, market_type, market_side, line_value, event_key, semantic_key, market_group
        FROM bets
        WHERE event_key=? AND result_status='pending'
        ORDER BY id DESC
    """, (parsed.get("event_key"),))


def get_last_bets(user_id: int, limit: int = 20):
    return all_rows("""
        SELECT id, sport, match_name, market, odds, stake, bookmaker, created_at, result_status
        FROM bets WHERE user_id=? ORDER BY id DESC LIMIT ?
    """, (user_id, limit))


def get_last_bet(user_id: int):
    rows = get_last_bets(user_id, 1)
    return rows[0] if rows else None


def get_bet_by_id(bet_id: int):
    return one("""
        SELECT id, sport, tournament, match_name, match_date, match_time, market, odds, ev, bookmaker, stake,
               created_at, result_status, payout, profit, shift_id
        FROM bets WHERE id=?
    """, (bet_id,))


def get_bets_by_shift(shift_id: int, offset: int = 0, limit: int = BET_PAGE_SIZE):
    rows = all_rows("""
        SELECT id, sport, match_name, market, odds, stake, bookmaker, created_at, result_status
        FROM bets WHERE shift_id=? ORDER BY id DESC LIMIT ? OFFSET ?
    """, (shift_id, limit, offset))
    total = one("SELECT COUNT(*) AS c FROM bets WHERE shift_id=?", (shift_id,))["c"]
    return rows, total


def get_shift_stats(shift_id: int):
    return one("""
        SELECT COUNT(*) AS total_bets, COALESCE(SUM(stake),0) AS total_stake, COALESCE(AVG(odds),0) AS avg_odds,
               COALESCE(AVG(ev),0) AS avg_ev,
               COALESCE(SUM(CASE WHEN result_status='win' THEN 1 ELSE 0 END),0) AS wins,
               COALESCE(SUM(CASE WHEN result_status='lose' THEN 1 ELSE 0 END),0) AS loses,
               COALESCE(SUM(CASE WHEN result_status='half_win' THEN 1 ELSE 0 END),0) AS half_wins,
               COALESCE(SUM(CASE WHEN result_status='half_lose' THEN 1 ELSE 0 END),0) AS half_loses,
               COALESCE(SUM(CASE WHEN result_status='refund' THEN 1 ELSE 0 END),0) AS refunds,
               COALESCE(SUM(CASE WHEN result_status='pending' THEN 1 ELSE 0 END),0) AS pendings,
               COALESCE(SUM(profit),0) AS total_profit
        FROM bets WHERE shift_id=?
    """, (shift_id,))


def get_market_stats_by_shift(shift_id: int):
    rows = all_rows("""
        SELECT COALESCE(market_group,'Другое') AS market_group, COUNT(*) AS count, COALESCE(SUM(stake),0) AS total_stake,
               COALESCE(AVG(stake),0) AS avg_stake, COALESCE(AVG(odds),0) AS avg_odds, COALESCE(AVG(ev),0) AS avg_ev,
               COALESCE(SUM(profit),0) AS profit,
               COALESCE(SUM(CASE WHEN result_status='pending' THEN 1 ELSE 0 END),0) AS pending
        FROM bets WHERE shift_id=?
        GROUP BY COALESCE(market_group,'Другое') ORDER BY COUNT(*) DESC
    """, (shift_id,))
    total = one("SELECT COUNT(*) AS c FROM bets WHERE shift_id=?", (shift_id,))["c"]
    return rows, total


def get_bookmaker_stats_by_shift(shift_id: int):
    rows = all_rows("""
        SELECT COALESCE(bookmaker,'Other') AS bookmaker, COUNT(*) AS count, COALESCE(SUM(stake),0) AS total_stake,
               COALESCE(AVG(stake),0) AS avg_stake, COALESCE(AVG(odds),0) AS avg_odds, COALESCE(SUM(profit),0) AS profit
        FROM bets WHERE shift_id=?
        GROUP BY COALESCE(bookmaker,'Other') ORDER BY COUNT(*) DESC
    """, (shift_id,))
    total = one("SELECT COUNT(*) AS c FROM bets WHERE shift_id=?", (shift_id,))["c"]
    return rows, total


def get_today_stats(user_id: int):
    start = now_dt().replace(hour=0, minute=0, second=0, microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    end = now_dt().replace(hour=23, minute=59, second=59, microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    return one("""
        SELECT COUNT(*) AS total_bets, COALESCE(SUM(stake),0) AS total_stake, COALESCE(AVG(odds),0) AS avg_odds,
               COALESCE(AVG(ev),0) AS avg_ev, COALESCE(SUM(profit),0) AS total_profit,
               COALESCE(SUM(CASE WHEN result_status='pending' THEN 1 ELSE 0 END),0) AS pendings
        FROM bets WHERE user_id=? AND created_at>=? AND created_at<=?
    """, (user_id, start, end))


def update_bet_result(bet_id: int, result_status: str):
    row = one("SELECT stake, odds FROM bets WHERE id=?", (bet_id,))
    if not row:
        return False
    payout, profit = calc_settlement(row["stake"], row["odds"], result_status)
    exec_sql("UPDATE bets SET result_status=?, payout=?, profit=? WHERE id=?", (result_status, payout, profit, bet_id))
    log_info(f"Bet result updated | bet_id={bet_id} | status={result_status}")
    return True


def update_bet_stake(bet_id: int, new_stake: float):
    row = one("SELECT shift_id, stake, odds, result_status FROM bets WHERE id=?", (bet_id,))
    if not row:
        return False, "Ставка не найдена."
    delta = new_stake - row["stake"]
    payout, profit = calc_settlement(new_stake, row["odds"], row["result_status"])
    conn = db()
    conn.execute("UPDATE bets SET stake=?, payout=?, profit=? WHERE id=?", (new_stake, payout, profit, bet_id))
    conn.execute("UPDATE shifts SET spent=spent+? WHERE id=?", (delta, row["shift_id"]))
    conn.commit()
    conn.close()
    log_info(f"Bet stake updated | bet_id={bet_id} | old={row['stake']} | new={new_stake}")
    return True, row["stake"]


def get_recent_logs(limit: int = 10):
    return all_rows("SELECT level, message, created_at FROM logs ORDER BY id DESC LIMIT ?", (limit,))


def get_due_reminders():
    rows = all_rows("""
        SELECT id, user_id, match_name, market, match_start_at, stake, odds, bookmaker
        FROM bets WHERE reminder_sent=0 AND result_status='pending' AND match_start_at IS NOT NULL
    """)
    current = now_dt()
    upper = current + timedelta(minutes=REMINDER_MINUTES)
    due = []
    for r in rows:
        try:
            dt = datetime.fromisoformat(r["match_start_at"])
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=TIMEZONE)
        except Exception:
            continue
        if current <= dt <= upper:
            due.append(dict(r) | {"dt": dt})
    return due


def mark_reminder_sent(bet_id: int):
    exec_sql("UPDATE bets SET reminder_sent=1 WHERE id=?", (bet_id,))


# =========================
# PARSER
# =========================
def parse_bet_diagnostics(text: str) -> str:
    checks = []
    if "⚽️🏒🎾" not in text:
        checks.append("не нашёл строку спорта/турнира с маркером ⚽️🏒🎾")
    if "🚩" not in text:
        checks.append("не нашёл строку матча с маркером 🚩")
    if "коэф" not in normalize_text(text):
        checks.append("не нашёл коэффициент: в тексте нет слова 'коэф.'")
    if not re.search(r"\d{1,2}:\d{2}\s+\d{2}/\d{2}", text):
        checks.append("не нашёл время и дату матча в формате 12:30 28/03")
    if "Ставка сделана" not in text:
        checks.append("не нашёл букмекера: строку 'Ставка сделана👉 ...'")
    if not checks:
        checks.append("структура похожа на ставку, но формат отличается от ожидаемого шаблона")
    return "\n".join(f"• {x}" for x in checks)


def parse_bet(text: str):
    pattern = re.compile(r"(⚽️🏒🎾.*?)(?=\n\s*⚽️🏒🎾|\Z)", re.S)
    matches = pattern.findall(text)
    if not matches:
        return None

    block = None
    for c in matches:
        if (
            re.search(r"⚽️🏒🎾\s*(.+?)\n", c, re.S)
            and re.search(r"🚩\s*(.+?),\s*(\d{1,2}:\d{2})\s+(\d{2}/\d{2})", c)
            and re.search(r"❗️\s*(.+?)\s*коэф\.?\s*([\d.,]+)❗️", c, re.S)
        ):
            block = c.strip()
            break

    if not block:
        return None

    sport_line = re.search(r"⚽️🏒🎾\s*(.+?)\n", block, re.S)
    event_line = re.search(r"🚩\s*(.+?),\s*(\d{1,2}:\d{2})\s+(\d{2}/\d{2})", block)
    market_line = re.search(r"❗️\s*(.+?)\s*коэф\.?\s*([\d.,]+)❗️", block, re.S)
    ev_line = re.search(r"Математическое ожидание\s*≈\s*([\d.,]+)%", block, re.I)
    bk_line = re.search(r"Ставка сделана👉\s*([^\n(]+)", block, re.I)

    if not sport_line or not event_line or not market_line:
        return None

    sport, tournament = split_sport_tournament(sport_line.group(1).strip())
    match_name = event_line.group(1).strip()
    match_time = event_line.group(2).strip()
    match_date = event_line.group(3).strip()
    market = re.sub(r"\s+", " ", market_line.group(1)).strip()
    odds = float(market_line.group(2).replace(",", "."))
    ev = float(ev_line.group(1).replace(",", ".")) if ev_line else None
    bookmaker = bk_line.group(1).strip() if bk_line else ""
    match_start_at = parse_match_start(match_date, match_time)

    parsed = {
        "sport": sport, "tournament": tournament, "match_name": match_name,
        "match_time": match_time, "match_date": match_date,
        "match_start_at": match_start_at.isoformat(),
        "market": market, "odds": odds, "ev": ev,
        "bookmaker": bookmaker, "hash": hash_text(block), "source_text": block,
    }
    parsed["event_key"] = event_key_from(match_name, parsed["match_start_at"])
    return normalize_market(parsed)


# =========================
# RISK ENGINE
# =========================
def compact_risk_text(parsed: dict) -> str:
    status = parsed.get("risk_status", "new")
    notes = parsed.get("risk_notes", "")

    if status == "new":
        return "━━━━━━━━━━━━━━\n✅ <b>РИСКИ: ЧИСТО</b>\nПовторов / коридоров / вилок среди pending-ставок не найдено.\n━━━━━━━━━━━━━━"

    title = {
        "duplicate": "⚠️ <b>ПОВТОР</b>",
        "corridor": "🟣 <b>КОРИДОР</b>",
        "arbitrage": "🟢 <b>ВИЛКА</b>",
        "mixed": "🔀 <b>НЕСКОЛЬКО РИСКОВ</b>",
        "opposite_no_value": "🟡 <b>ПРОТИВОПОЛОЖНОЕ ПЛЕЧО БЕЗ ПЛЮСА</b>",
    }.get(status, f"⚠️ <b>{status}</b>")

    important = []
    for line in notes.splitlines():
        clean = line.strip()
        if not clean:
            continue
        if any(x in clean.lower() for x in [
            "уже есть", "новая", "коридор", "ширина", "рекоменд", "потенциаль",
            "прибыль", "результат", "вероятност", "та же сторона", "усиление",
            "вне коридора"
        ]):
            important.append(clean)
        if len(important) >= 8:
            break
    if not important:
        important = notes.splitlines()[:7]

    return "━━━━━━━━━━━━━━\n" + title + "\n\n" + "\n".join(important) + "\n━━━━━━━━━━━━━━"


def analyze_risk(parsed: dict):
    rows = get_pending_similar_bets(parsed)
    notes = []
    statuses = []

    for row in rows:
        existing = dict(row)
        same_type = existing["market_type"] == parsed.get("market_type")
        same_selection = existing["selection_side"] == parsed.get("selection_side") and parsed.get("selection_side") not in {None, "unknown"}

        if same_type and same_selection and parsed.get("market_type") in {"handicap", "moneyline"}:
            statuses.append("duplicate")
            notes.append(
                f"Уже есть: {existing['market']}\n"
                f"🏦 {bookmaker_label(existing['bookmaker'])} | 📈 {existing['odds']} | 💸 {existing['stake']}\n"
                f"Новая: {parsed['market']}\n"
                f"Это та же сторона и тот же тип рынка. Похоже на усиление позиции."
            )
            continue

        if same_type and parsed.get("market_type") == "total" and existing.get("market_side") == parsed.get("market_side"):
            statuses.append("duplicate")
            notes.append(
                f"Уже есть: {existing['market']} | 📈 {existing['odds']} | 💸 {existing['stake']}\n"
                f"Новая: {parsed['market']}\n"
                f"Это тот же тип тотала в ту же сторону."
            )
            continue

        if same_type and parsed.get("market_type") == "handicap" and existing.get("selection_side") in {"home", "away"} and parsed.get("selection_side") in {"home", "away"} and existing.get("selection_side") != parsed.get("selection_side"):
            line1, line2 = existing.get("line_value"), parsed.get("line_value")
            if line1 is None or line2 is None:
                continue
            cond1, val1 = handicap_bound(existing["selection_side"], float(line1))
            cond2, val2 = handicap_bound(parsed["selection_side"], float(line2))
            if cond1 and cond2:
                lower = max([v for c, v in [(cond1, val1), (cond2, val2)] if c == "gt"], default=None)
                upper = min([v for c, v in [(cond1, val1), (cond2, val2)] if c == "lt"], default=None)
                stake2, payout, total_stake, hedge_profit, hedge_roi, implied = arbitrage_metrics(existing["stake"], existing["odds"], parsed["odds"])

                if lower is not None and upper is not None and lower < upper:
                    width = round(upper - lower, 2)
                    corridor_profit = round(existing["stake"] * (existing["odds"] - 1) + stake2 * (parsed["odds"] - 1), 2)
                    statuses.append("corridor")
                    notes.append(
                        f"Уже есть: {existing['market']} | 📈 {existing['odds']} | 💸 {existing['stake']}\n"
                        f"Новая: {parsed['market']} | 📈 {parsed['odds']}\n"
                        f"Коридор по марже 1-й команды: <b>{lower} — {upper}</b>\n"
                        f"Ширина коридора: <b>{width}</b> очков\n"
                        f"Рекомендованная сумма второго плеча: <b>{stake2}</b>\n"
                        f"При попадании в коридор: примерно <b>{corridor_profit}</b>\n"
                        f"Вне коридора при балансировке: примерно <b>{hedge_profit}</b> ({hedge_roi}%)"
                    )
                else:
                    statuses.append("arbitrage" if implied < 1 else "opposite_no_value")
                    notes.append(
                        f"Уже есть: {existing['market']} | 📈 {existing['odds']} | 💸 {existing['stake']}\n"
                        f"Новая: {parsed['market']} | 📈 {parsed['odds']}\n"
                        f"Сумма вероятностей: <b>{implied}</b>\n"
                        f"Рекомендованная сумма второго плеча: <b>{stake2}</b>\n"
                        f"Потенциальный результат при балансировке: <b>{hedge_profit}</b> ({hedge_roi}%)"
                    )

        if same_type and parsed.get("market_type") == "total" and existing.get("market_side") != parsed.get("market_side"):
            line1, line2 = existing.get("line_value"), parsed.get("line_value")
            if line1 is None or line2 is None:
                continue
            over_line = line1 if existing["market_side"] == "over" else line2 if parsed["market_side"] == "over" else None
            under_line = line1 if existing["market_side"] == "under" else line2 if parsed["market_side"] == "under" else None
            stake2, payout, total_stake, hedge_profit, hedge_roi, implied = arbitrage_metrics(existing["stake"], existing["odds"], parsed["odds"])

            if over_line is not None and under_line is not None and over_line < under_line:
                width = round(under_line - over_line, 2)
                corridor_profit = round(existing["stake"] * (existing["odds"] - 1) + stake2 * (parsed["odds"] - 1), 2)
                statuses.append("corridor")
                notes.append(
                    f"Уже есть: {existing['market']} | 📈 {existing['odds']} | 💸 {existing['stake']}\n"
                    f"Новая: {parsed['market']} | 📈 {parsed['odds']}\n"
                    f"Коридор: итоговый тотал между <b>{over_line}</b> и <b>{under_line}</b>\n"
                    f"Ширина: <b>{width}</b>\n"
                    f"Рекомендованная сумма второго плеча: <b>{stake2}</b>\n"
                    f"При попадании в коридор: примерно <b>{corridor_profit}</b>"
                )
            else:
                statuses.append("arbitrage" if implied < 1 else "opposite_no_value")
                notes.append(
                    f"Уже есть: {existing['market']} | 📈 {existing['odds']} | 💸 {existing['stake']}\n"
                    f"Новая: {parsed['market']} | 📈 {parsed['odds']}\n"
                    f"Сумма вероятностей: <b>{implied}</b>\n"
                    f"Рекомендованная сумма второго плеча: <b>{stake2}</b>\n"
                    f"Потенциальный результат при балансировке: <b>{hedge_profit}</b> ({hedge_roi}%)"
                )

    if not statuses:
        status, notes_text = "new", "✅ Повторов, коридоров и вилок среди pending-ставок не найдено."
    elif "corridor" in statuses and "arbitrage" in statuses:
        status, notes_text = "mixed", "\n\n".join(notes)
    elif "corridor" in statuses:
        status, notes_text = "corridor", "\n\n".join(notes)
    elif "arbitrage" in statuses:
        status, notes_text = "arbitrage", "\n\n".join(notes)
    elif "duplicate" in statuses:
        status, notes_text = "duplicate", "\n\n".join(notes)
    else:
        status, notes_text = "opposite_no_value", "\n\n".join(notes)

    parsed["risk_status"] = status
    parsed["risk_notes"] = notes_text
    return parsed


# =========================
# EXPORTS
# =========================
def export_bets_to_csv(user_id: int) -> str | None:
    rows = all_rows("""
        SELECT id, shift_id, created_at, sport, tournament, match_name,
               match_date, match_time, match_start_at, market, market_group,
               odds, ev, bookmaker, stake, result_status, payout, profit,
               risk_status, risk_notes
        FROM bets WHERE user_id = ?
        ORDER BY id DESC
    """, (user_id,))
    if not rows:
        return None

    filename = f"bets_export_all_{now_dt().strftime('%Y%m%d_%H%M%S')}.csv"
    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(rows[0].keys())
        for r in rows:
            writer.writerow(list(r))
    return filename


def export_shift_to_xlsx(user_id: int, shift_id: int) -> str | None:
    shift = get_shift_by_id(shift_id, user_id)
    if not shift:
        return None

    bets = all_rows("""
        SELECT id, created_at, sport, tournament, match_name, match_date,
               match_time, match_start_at, market, market_group, odds, ev,
               bookmaker, stake, result_status, payout, profit, risk_status, risk_notes
        FROM bets WHERE shift_id = ?
        ORDER BY id ASC
    """, (shift_id,))
    rejected = all_rows("""
        SELECT created_at, match_name, market, odds, bookmaker, risk_status, risk_notes, reason
        FROM rejected_bets WHERE user_id = ?
        ORDER BY id DESC
    """, (user_id,))

    if not bets:
        return None

    filename = f"shift_{shift_id}_report_{now_dt().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=13)

    stats = get_shift_stats(shift_id)
    roi = calc_roi(stats["total_profit"], stats["total_stake"])

    ws["A1"] = "Shift Summary"
    ws["A1"].font = title_font
    summary_rows = [
        ["Shift ID", shift["id"]],
        ["Started", shift["started_at"]],
        ["Ended", shift["ended_at"] or "active"],
        ["Budget", shift["budget"]],
        ["Spent", shift["spent"]],
        ["Bets", stats["total_bets"]],
        ["Total stake", round(stats["total_stake"], 2)],
        ["Avg odds", round(stats["avg_odds"], 2)],
        ["Avg EV", round(stats["avg_ev"], 2)],
        ["Profit", round(stats["total_profit"], 2)],
        ["ROI %", roi],
        ["Win", stats["wins"]],
        ["Lose", stats["loses"]],
        ["Half win", stats["half_wins"]],
        ["Half lose", stats["half_loses"]],
        ["Refund", stats["refunds"]],
        ["Pending", stats["pendings"]],
    ]
    for row in summary_rows:
        ws.append(row)

    ws.append([])
    ws.append(["Market", "Count", "Share %", "Total stake", "Avg stake", "Avg odds", "Avg EV", "Profit", "ROI %", "Pending"])
    for c in ws[ws.max_row]:
        c.font = header_font; c.fill = header_fill

    market_rows, total = get_market_stats_by_shift(shift_id)
    for r in market_rows:
        ws.append([
            r["market_group"], r["count"], round(r["count"] / total * 100, 2) if total else 0,
            round(r["total_stake"], 2), round(r["avg_stake"], 2), round(r["avg_odds"], 2),
            round(r["avg_ev"], 2), round(r["profit"], 2), calc_roi(r["profit"], r["total_stake"]),
            r["pending"],
        ])

    ws.append([])
    ws.append(["Bookmaker", "Count", "Share %", "Total stake", "Avg stake", "Avg odds", "Profit", "ROI %"])
    for c in ws[ws.max_row]:
        c.font = header_font; c.fill = header_fill

    bk_rows, total_bk = get_bookmaker_stats_by_shift(shift_id)
    for r in bk_rows:
        ws.append([
            bookmaker_label(r["bookmaker"]), r["count"], round(r["count"] / total_bk * 100, 2) if total_bk else 0,
            round(r["total_stake"], 2), round(r["avg_stake"], 2), round(r["avg_odds"], 2),
            round(r["profit"], 2), calc_roi(r["profit"], r["total_stake"]),
        ])

    ws_bets = wb.create_sheet("Bets")
    ws_bets.append(bets[0].keys())
    for c in ws_bets[1]:
        c.font = header_font; c.fill = header_fill
    for r in bets:
        row = list(r)
        row[12] = bookmaker_label(row[12])
        ws_bets.append(row)

    ws_risk = wb.create_sheet("Risks_Rejected")
    ws_risk.append(["created_at", "match_name", "market", "odds", "bookmaker", "risk_status", "risk_notes", "reason"])
    for c in ws_risk[1]:
        c.font = header_font; c.fill = header_fill
    for r in rejected:
        row = list(r)
        row[4] = bookmaker_label(row[4])
        ws_risk.append(row)

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = 0
            letter = col[0].column_letter
            for cell in col:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.value is not None:
                    width = max(width, len(str(cell.value)))
            sheet.column_dimensions[letter].width = min(width + 2, 60)

    wb.save(filename)
    return filename


# =========================
# KEYBOARDS
# =========================
def main_menu_kb():
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="🎯 Смена"), KeyboardButton(text="📚 Ставки")],
        [KeyboardButton(text="📊 Статистика"), KeyboardButton(text="⚙️ Сервис")],
        [KeyboardButton(text="❌ Отмена")],
    ], resize_keyboard=True, is_persistent=True)


def shift_menu_kb(active: bool):
    keyboard = [[KeyboardButton(text="🚀 Начать смену")]] if not active else [[KeyboardButton(text="📍 Текущая смена"), KeyboardButton(text="🏁 Завершить смену")]]
    keyboard.append([KeyboardButton(text="⬅️ Назад")])
    return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True, is_persistent=True)


def bets_menu_kb():
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="➕ Добавить ставку")],
        [KeyboardButton(text="📂 Выбрать смену")],
        [KeyboardButton(text="🧾 Последняя ставка"), KeyboardButton(text="📚 Последние 20 ставок")],
        [KeyboardButton(text="⬅️ Назад")],
    ], resize_keyboard=True, is_persistent=True)


def stats_menu_kb():
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="📈 Статистика по смене"), KeyboardButton(text="📋 Список смен")],
        [KeyboardButton(text="📦 XLSX текущей смены")],
        [KeyboardButton(text="📅 Статистика за день"), KeyboardButton(text="📌 Ближайшие матчи")],
        [KeyboardButton(text="📤 Export CSV all")],
        [KeyboardButton(text="⬅️ Назад")],
    ], resize_keyboard=True, is_persistent=True)


def service_menu_kb():
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="📋 Логи")], [KeyboardButton(text="⬅️ Назад")]], resize_keyboard=True, is_persistent=True)


def yes_no_kb():
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="✅ Подтвердить"), KeyboardButton(text="❌ Отмена")]], resize_keyboard=True, is_persistent=True)


def risk_decision_kb():
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="💾 Сохранить"), KeyboardButton(text="🚫 Отказаться")], [KeyboardButton(text="❌ Отмена")]], resize_keyboard=True, is_persistent=True)


def amount_retry_kb():
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="🔁 Повторить ввод суммы")], [KeyboardButton(text="❌ Отмена")]], resize_keyboard=True, is_persistent=True)


def result_kb():
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="🕒 В ожидании")],
        [KeyboardButton(text="✅ Выигрыш"), KeyboardButton(text="❌ Проигрыш")],
        [KeyboardButton(text="🟡 Половина выигрыша"), KeyboardButton(text="🟠 Половина проигрыша")],
        [KeyboardButton(text="↩️ Возврат")],
        [KeyboardButton(text="❌ Отмена")],
    ], resize_keyboard=True, is_persistent=True)


def shift_list_kb(has_prev: bool, has_next: bool):
    row = []
    if has_prev: row.append(KeyboardButton(text="⬅️ Пред. смены"))
    if has_next: row.append(KeyboardButton(text="➡️ След. смены"))
    keyboard = [row] if row else []
    keyboard.append([KeyboardButton(text="⬅️ Назад")])
    return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True, is_persistent=True)


def selected_shift_kb():
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="📈 Отчет смены"), KeyboardButton(text="📚 Ставки смены")],
        [KeyboardButton(text="📦 XLSX смены")],
        [KeyboardButton(text="⬅️ Назад")],
    ], resize_keyboard=True, is_persistent=True)


def selected_bet_action_kb():
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="✏️ Изменить сумму выбранной")],
        [KeyboardButton(text="🏷 Рассчитать выбранную")],
        [KeyboardButton(text="⬅️ Назад")],
    ], resize_keyboard=True, is_persistent=True)


# =========================
# FORMATTERS
# =========================
def format_shift_stats_text(shift_id: int, title: str = "📈 <b>Статистика по смене</b>") -> str:
    shift = get_shift_by_id(shift_id, OWNER_ID)
    if not shift:
        return "📭 Смена не найдена."
    s = get_shift_stats(shift_id)
    remain = round(shift["budget"] - shift["spent"], 2)
    roi = calc_roi(s["total_profit"], s["total_stake"])

    text = (
        f"{title}\n━━━━━━━━━━━━━━\n"
        f"🆔 Смена: <b>{shift_id}</b>\n"
        f"🕒 Начало: <b>{shift['started_at']} МСК</b>\n"
        f"🏁 Конец: <b>{shift['ended_at'] or 'активна'}</b>\n\n"
        f"🎯 Ставок: <b>{s['total_bets']}</b>\n"
        f"💸 Общая сумма: <b>{round(s['total_stake'], 2)}</b>\n"
        f"📈 Средний КФ: <b>{round(s['avg_odds'], 2) if s['total_bets'] else 0}</b>\n"
        f"🧠 Среднее EV: <b>{round(s['avg_ev'], 2) if s['total_bets'] else 0}</b>\n\n"
        f"✅ Выигрыш: <b>{s['wins']}</b>\n"
        f"❌ Проигрыш: <b>{s['loses']}</b>\n"
        f"🟡 Half win: <b>{s['half_wins']}</b>\n"
        f"🟠 Half lose: <b>{s['half_loses']}</b>\n"
        f"↩️ Возврат: <b>{s['refunds']}</b>\n"
        f"🕒 Pending: <b>{s['pendings']}</b>\n\n"
        f"💰 Бюджет: <b>{shift['budget']}</b>\n"
        f"💸 Поставлено: <b>{shift['spent']}</b>\n"
        f"🟢 Остаток: <b>{remain}</b>\n"
        f"📊 Прибыль: <b>{round(s['total_profit'], 2)}</b>\n"
        f"📐 ROI: <b>{roi}%</b>"
    )

    market_rows, total = get_market_stats_by_shift(shift_id)
    if market_rows:
        text += "\n\n📌 <b>Маркеты</b>"
        for r in market_rows[:10]:
            share = round(r["count"] / total * 100, 2) if total else 0
            text += (
                f"\n• <b>{r['market_group']}</b>: {r['count']} ставок, {share}%, "
                f"ср. сумма {round(r['avg_stake'], 2)}, ср. КФ {round(r['avg_odds'], 2)}, "
                f"ROI {calc_roi(r['profit'], r['total_stake'])}%"
            )
    return text


def format_bet_card(bet) -> str:
    return (
        "🎯 <b>Выбрана ставка</b>\n━━━━━━━━━━━━━━\n"
        f"🆔 ID: <b>{bet['id']}</b>\n"
        f"🏟 <b>{bet['match_name']}</b>\n\n"
        f"📌 {bet['market']}\n\n"
        f"💸 Сумма: <b>{bet['stake']}</b>\n"
        f"📈 КФ: <b>{bet['odds']}</b>\n"
        f"🏦 {bookmaker_label(bet['bookmaker'])}\n"
        f"🏷 Сейчас: <b>{RESULT_LABELS.get(bet['result_status'], bet['result_status'])}</b>"
    )


# =========================
# NAVIGATION / COMMANDS
# =========================
@dp.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID:
        await message.answer("⛔ Этот бот доступен только владельцу.")
        return
    await state.clear()
    await message.answer("🚀 <b>Бот учёта ставок + риск-сканер запущен</b>\n\nВыбери раздел 👇", reply_markup=main_menu_kb())


@dp.message(F.text == "⬅️ Назад")
async def back_to_main(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("🏠 Главное меню", reply_markup=main_menu_kb())


@dp.message(F.text == "❌ Отмена")
async def cancel_action(message: Message, state: FSMContext):
    data = await state.get_data()
    parsed = data.get("pending_bet")
    if parsed:
        save_rejected_bet(message.from_user.id, parsed, "cancelled")
    await state.clear()
    await message.answer("❌ Действие отменено.", reply_markup=main_menu_kb())


@dp.message(F.text == "🎯 Смена")
async def open_shift_menu(message: Message, state: FSMContext):
    await state.clear()
    active = get_active_shift(message.from_user.id)
    if active:
        await message.answer(
            f"🟢 <b>Смена активна</b>\n\n"
            f"🆔 ID: <b>{active['id']}</b>\n"
            f"🕒 {active['started_at']} МСК\n"
            f"💰 Бюджет: <b>{active['budget']}</b>\n"
            f"💸 Поставлено: <b>{active['spent']}</b>\n"
            f"🟢 Остаток: <b>{round(active['budget'] - active['spent'], 2)}</b>",
            reply_markup=shift_menu_kb(True),
        )
    else:
        await message.answer("🎯 <b>Раздел «Смена»</b>", reply_markup=shift_menu_kb(False))


@dp.message(F.text == "📚 Ставки")
async def open_bets_menu(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("📚 <b>Раздел «Ставки»</b>", reply_markup=bets_menu_kb())


@dp.message(F.text == "📊 Статистика")
async def open_stats_menu(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("📊 <b>Раздел «Статистика»</b>", reply_markup=stats_menu_kb())


@dp.message(F.text == "⚙️ Сервис")
async def open_service_menu(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("⚙️ <b>Раздел «Сервис»</b>", reply_markup=service_menu_kb())


# =========================
# SHIFT FLOW
# =========================
@dp.message(F.text == "🚀 Начать смену")
async def start_shift_button(message: Message, state: FSMContext):
    if has_recent_action(message.from_user.id, "start_shift"):
        return
    if get_active_shift(message.from_user.id):
        await message.answer("ℹ️ Смена уже активна.", reply_markup=shift_menu_kb(True))
        return
    await state.set_state(S.waiting_budget)
    await message.answer(
        "💰 <b>Введи бюджет смены</b>\n\nПример: <code>10000</code>",
        reply_markup=ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="❌ Отмена")]], resize_keyboard=True, is_persistent=True),
    )


@dp.message(S.waiting_budget)
async def budget_input(message: Message, state: FSMContext):
    try:
        budget = as_float(message.text.strip())
        if budget <= 0:
            raise ValueError
    except Exception:
        await message.answer("⚠️ Бюджет не распознан. Введи число, например: <code>10000</code>")
        return
    start_shift_db(message.from_user.id, budget)
    await state.clear()
    await message.answer(f"✅ <b>Смена начата</b>\n💰 Бюджет: <b>{budget}</b>", reply_markup=shift_menu_kb(True))


@dp.message(F.text == "📍 Текущая смена")
async def current_shift_handler(message: Message):
    active = get_active_shift(message.from_user.id)
    if not active:
        await message.answer("📭 Активной смены нет.", reply_markup=shift_menu_kb(False))
        return
    await message.answer(format_shift_stats_text(active["id"], "📊 <b>Текущая смена</b>"), reply_markup=shift_menu_kb(True))


@dp.message(F.text == "🏁 Завершить смену")
async def end_shift_handler(message: Message, state: FSMContext):
    active = get_active_shift(message.from_user.id)
    if not active:
        await message.answer("📭 Активной смены нет.", reply_markup=shift_menu_kb(False))
        return
    await state.set_state(S.waiting_end_shift_confirm)
    await message.answer(
        f"🏁 <b>Подтвердить завершение смены?</b>\n\n"
        f"💰 Бюджет: <b>{active['budget']}</b>\n"
        f"💸 Поставлено: <b>{active['spent']}</b>\n"
        f"🟢 Остаток: <b>{round(active['budget'] - active['spent'], 2)}</b>",
        reply_markup=yes_no_kb(),
    )


@dp.message(S.waiting_end_shift_confirm, F.text == "✅ Подтвердить")
async def confirm_end_shift(message: Message, state: FSMContext):
    active = get_active_shift(message.from_user.id)
    if not active:
        await state.clear()
        await message.answer("📭 Активной смены уже нет.", reply_markup=shift_menu_kb(False))
        return
    shift_id = active["id"]
    end_shift_db(shift_id)
    await state.clear()
    await message.answer(f"🏁 <b>Смена завершена</b>\n\n{format_shift_stats_text(shift_id, '📊 <b>Итог смены</b>')}", reply_markup=shift_menu_kb(False))


# =========================
# ADD BET FLOW
# =========================
@dp.message(F.text == "➕ Добавить ставку")
async def add_bet_hint(message: Message):
    if not get_active_shift(message.from_user.id):
        await message.answer("⚠️ Сначала начни смену.", reply_markup=shift_menu_kb(False))
        return
    await message.answer(
        "📥 <b>Перешли мне сообщение со ставкой</b>\n\n"
        "Я проверю: повтор / коридор / вилку.\n"
        "После проверки ты решишь: сохранить или отказаться.",
        reply_markup=bets_menu_kb(),
    )


@dp.message(F.text == "💾 Сохранить")
async def risk_save_handler(message: Message, state: FSMContext):
    if await state.get_state() != S.waiting_risk_decision.state:
        return
    await state.set_state(S.waiting_bet_amount)
    await message.answer("💬 Теперь напиши сумму ставки.", reply_markup=amount_retry_kb())


@dp.message(F.text == "🚫 Отказаться")
async def risk_reject_handler(message: Message, state: FSMContext):
    data = await state.get_data()
    parsed = data.get("pending_bet")
    if parsed:
        save_rejected_bet(message.from_user.id, parsed, "manual_reject")
    await state.clear()
    await message.answer("🚫 Ставка отклонена и сохранена в rejected_bets.", reply_markup=bets_menu_kb())


@dp.message(F.text == "🔁 Повторить ввод суммы")
async def retry_amount_handler(message: Message, state: FSMContext):
    if await state.get_state() != S.waiting_bet_amount.state:
        await message.answer("ℹ️ Сейчас нет активного ввода суммы.", reply_markup=bets_menu_kb())
        return
    await message.answer("🔁 Напиши сумму заново. Пример: <code>1500</code>", reply_markup=amount_retry_kb())


# =========================
# QUICK BETS VIEW
# =========================
@dp.message(F.text == "🧾 Последняя ставка")
async def last_bet_handler(message: Message):
    row = get_last_bet(message.from_user.id)
    if not row:
        await message.answer("📭 Пока нет ставок.", reply_markup=bets_menu_kb())
        return
    await message.answer(
        "🧾 <b>Последняя ставка</b>\n━━━━━━━━━━━━━━\n"
        f"🏅 {row['sport']}\n"
        f"🏟 <b>{row['match_name']}</b>\n\n"
        f"📌 {row['market']}\n\n"
        f"📈 КФ: <b>{row['odds']}</b>\n"
        f"💸 Сумма: <b>{row['stake']}</b>\n"
        f"🏦 {bookmaker_label(row['bookmaker'])}\n"
        f"🏷 {RESULT_LABELS.get(row['result_status'], row['result_status'])}",
        reply_markup=bets_menu_kb(),
    )


@dp.message(F.text == "📚 Последние 20 ставок")
async def last_20_handler(message: Message):
    rows = get_last_bets(message.from_user.id, 20)
    if not rows:
        await message.answer("📭 Пока нет ставок.", reply_markup=bets_menu_kb())
        return
    lines = ["📚 <b>Последние 20 ставок</b>\n"]
    for i, r in enumerate(rows, 1):
        lines.append(
            f"{i}. <b>{r['sport']}</b> | {r['match_name']}\n"
            f"📌 {r['market']}\n"
            f"📈 {r['odds']} | 💸 {r['stake']} | 🏦 {bookmaker_label(r['bookmaker'])}\n"
            f"🏷 {RESULT_LABELS.get(r['result_status'], r['result_status'])}\n"
        )
    await message.answer("\n".join(lines), reply_markup=bets_menu_kb())


# =========================
# SHIFT SELECTION
# =========================
@dp.message(F.text.in_({"📋 Список смен", "📂 Выбрать смену"}))
async def shift_list_start(message: Message, state: FSMContext):
    await show_shift_page(message, state, 0)


async def show_shift_page(message: Message, state: FSMContext, page: int):
    offset = page * SHIFT_PAGE_SIZE
    rows, total = list_shifts(message.from_user.id, offset, SHIFT_PAGE_SIZE)
    if not rows:
        await message.answer("📭 Смен пока нет.", reply_markup=stats_menu_kb())
        return
    mapping = {}
    lines = [f"📋 <b>Смены {offset + 1}–{offset + len(rows)} из {total}</b>\n"]
    for idx, r in enumerate(rows, 1):
        s = get_shift_stats(r["id"])
        mapping[str(idx)] = r["id"]
        lines.append(
            f"{idx}. ID <b>{r['id']}</b> | {r['started_at']}\n"
            f"Статус: <b>{r['status']}</b>\n"
            f"Бюджет: <b>{r['budget']}</b> | Поставлено: <b>{r['spent']}</b>\n"
            f"Ставок: <b>{s['total_bets']}</b> | ROI: <b>{calc_roi(s['total_profit'], s['total_stake'])}%</b>\n"
        )
    await state.update_data(shift_page=page, shift_choices=mapping)
    await state.set_state(S.waiting_shift_number)
    await message.answer("\n".join(lines) + "\nНапиши номер смены.", reply_markup=shift_list_kb(page > 0, offset + len(rows) < total))


@dp.message(S.waiting_shift_number, F.text.in_({"➡️ След. смены", "⬅️ Пред. смены"}))
async def shift_page_nav(message: Message, state: FSMContext):
    data = await state.get_data()
    page = data.get("shift_page", 0)
    page = page + 1 if message.text == "➡️ След. смены" else max(0, page - 1)
    await show_shift_page(message, state, page)


@dp.message(S.waiting_shift_number)
async def shift_choose_number(message: Message, state: FSMContext):
    data = await state.get_data()
    mapping = data.get("shift_choices", {})
    text = (message.text or "").strip()
    if text not in mapping:
        await message.answer("⚠️ Номер смены не найден.")
        return
    shift_id = mapping[text]
    await state.update_data(selected_shift_id=shift_id)
    await state.set_state(None)
    await message.answer(format_shift_stats_text(shift_id, f"📊 <b>Выбрана смена {shift_id}</b>"), reply_markup=selected_shift_kb())


@dp.message(F.text == "📈 Отчет смены")
async def selected_shift_report(message: Message, state: FSMContext):
    data = await state.get_data()
    shift_id = data.get("selected_shift_id")
    if not shift_id:
        await message.answer("⚠️ Сначала выбери смену через 📋 Список смен.", reply_markup=stats_menu_kb())
        return
    await message.answer(format_shift_stats_text(shift_id), reply_markup=selected_shift_kb())


@dp.message(F.text == "📚 Ставки смены")
async def selected_shift_bets(message: Message, state: FSMContext):
    data = await state.get_data()
    shift_id = data.get("selected_shift_id")
    if not shift_id:
        await message.answer("⚠️ Сначала выбери смену через 📋 Список смен.", reply_markup=stats_menu_kb())
        return
    rows, total = get_bets_by_shift(shift_id, 0, BET_PAGE_SIZE)
    if not rows:
        await message.answer("📭 В этой смене нет ставок.", reply_markup=selected_shift_kb())
        return
    mapping = {}
    lines = [f"📚 <b>Ставки смены {shift_id}</b>\nПоказано 1–{len(rows)} из {total}\n"]
    for idx, r in enumerate(rows, 1):
        mapping[str(idx)] = r["id"]
        lines.append(
            f"{idx}. <b>{r['match_name']}</b>\n"
            f"📌 {r['market']}\n"
            f"💸 {r['stake']} | 📈 {r['odds']} | 🏦 {bookmaker_label(r['bookmaker'])}\n"
            f"🏷 {RESULT_LABELS.get(r['result_status'], r['result_status'])}\n"
        )
    await state.update_data(shift_bet_choices=mapping, selected_shift_id=shift_id)
    await state.set_state(S.waiting_shift_bet_number)
    await message.answer(
        "\n".join(lines) + "\nНапиши номер ставки.",
        reply_markup=ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="❌ Отмена")]], resize_keyboard=True, is_persistent=True),
    )


@dp.message(S.waiting_shift_bet_number)
async def selected_shift_bet_number(message: Message, state: FSMContext):
    text = (message.text or "").strip()
    data = await state.get_data()
    mapping = data.get("shift_bet_choices", {})
    if text not in mapping:
        await message.answer("⚠️ Номер ставки не найден. Напиши номер из списка.")
        return
    bet = get_bet_by_id(mapping[text])
    if not bet:
        await state.clear()
        await message.answer("⚠️ Не удалось найти ставку.", reply_markup=selected_shift_kb())
        return
    await state.update_data(selected_shift_bet_id=bet["id"], selected_shift_id=bet["shift_id"])
    await state.set_state(S.waiting_selected_bet_action)
    await message.answer(format_bet_card(bet), reply_markup=selected_bet_action_kb())


@dp.message(S.waiting_selected_bet_action, F.text == "✏️ Изменить сумму выбранной")
async def selected_bet_edit_amount_start(message: Message, state: FSMContext):
    data = await state.get_data()
    if not data.get("selected_shift_bet_id"):
        await state.clear()
        await message.answer("⚠️ Ставка не выбрана.", reply_markup=main_menu_kb())
        return
    await state.set_state(S.waiting_selected_bet_new_stake)
    await message.answer("✏️ Напиши новую сумму ставки.", reply_markup=ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text="❌ Отмена")]], resize_keyboard=True, is_persistent=True))


@dp.message(S.waiting_selected_bet_new_stake)
async def selected_bet_edit_amount_finish(message: Message, state: FSMContext):
    data = await state.get_data()
    bet_id = data.get("selected_shift_bet_id")
    shift_id = data.get("selected_shift_id")
    try:
        new_stake = as_float(message.text.strip())
        if new_stake <= 0:
            raise ValueError
    except Exception:
        await message.answer("⚠️ Сумма не распознана. Введи число, например: <code>1500</code>")
        return
    ok, old_stake = update_bet_stake(bet_id, new_stake)
    await state.update_data(selected_shift_id=shift_id)
    await state.set_state(None)
    await message.answer(
        f"✅ <b>Сумма обновлена</b>\n\nСтарая: <b>{old_stake}</b>\nНовая: <b>{new_stake}</b>" if ok else f"⚠️ {old_stake}",
        reply_markup=selected_shift_kb(),
    )


@dp.message(S.waiting_selected_bet_action, F.text == "🏷 Рассчитать выбранную")
async def selected_bet_result_start(message: Message, state: FSMContext):
    data = await state.get_data()
    if not data.get("selected_shift_bet_id"):
        await state.clear()
        await message.answer("⚠️ Ставка не выбрана.", reply_markup=main_menu_kb())
        return
    await state.set_state(S.waiting_selected_bet_result_status)
    await message.answer("🏷 Выбери результат:", reply_markup=result_kb())


@dp.message(S.waiting_selected_bet_result_status, F.text.in_(list(RESULT_MAP.keys())))
async def selected_bet_result_finish(message: Message, state: FSMContext):
    data = await state.get_data()
    bet_id = data.get("selected_shift_bet_id")
    shift_id = data.get("selected_shift_id")
    result_status = RESULT_MAP[message.text]
    if not bet_id:
        await state.clear()
        await message.answer("⚠️ Ставка не выбрана.", reply_markup=main_menu_kb())
        return
    ok = update_bet_result(bet_id, result_status)
    await state.update_data(selected_shift_id=shift_id)
    await state.set_state(None)
    await message.answer(
        f"✅ Результат обновлён: <b>{RESULT_LABELS[result_status]}</b>" if ok else "⚠️ Не удалось обновить результат.",
        reply_markup=selected_shift_kb(),
    )


# =========================
# STATS / EXPORTS
# =========================
@dp.message(F.text == "📈 Статистика по смене")
async def shift_stats_handler(message: Message):
    active = get_active_shift(message.from_user.id)
    if not active:
        await message.answer("📭 Активной смены нет.", reply_markup=stats_menu_kb())
        return
    await message.answer(format_shift_stats_text(active["id"]), reply_markup=stats_menu_kb())


@dp.message(F.text == "📅 Статистика за день")
async def today_stats_handler(message: Message):
    s = get_today_stats(message.from_user.id)
    await message.answer(
        f"📅 <b>Статистика за сегодня</b>\n━━━━━━━━━━━━━━\n"
        f"🎯 Ставок: <b>{s['total_bets']}</b>\n"
        f"💸 Сумма: <b>{round(s['total_stake'], 2)}</b>\n"
        f"📈 Ср.КФ: <b>{round(s['avg_odds'], 2) if s['total_bets'] else 0}</b>\n"
        f"📊 Прибыль: <b>{round(s['total_profit'], 2)}</b>\n"
        f"📐 ROI: <b>{calc_roi(s['total_profit'], s['total_stake'])}%</b>\n"
        f"🕒 Pending: <b>{s['pendings']}</b>",
        reply_markup=stats_menu_kb(),
    )


@dp.message(F.text == "📤 Export CSV all")
async def export_csv_all(message: Message):
    path = export_bets_to_csv(message.from_user.id)
    if not path:
        await message.answer("📭 Нет данных.", reply_markup=stats_menu_kb())
        return
    await message.answer_document(FSInputFile(path), caption="📤 CSV за всё время готов.", reply_markup=stats_menu_kb())


@dp.message(F.text == "📦 XLSX текущей смены")
async def export_active_shift_xlsx(message: Message):
    active = get_active_shift(message.from_user.id)
    if not active:
        await message.answer("📭 Активной смены нет. Для отчёта по старой смене открой 📋 Список смен.", reply_markup=stats_menu_kb())
        return
    path = export_shift_to_xlsx(message.from_user.id, active["id"])
    if not path:
        await message.answer("📭 В текущей смене нет ставок для XLSX.", reply_markup=stats_menu_kb())
        return
    await message.answer_document(FSInputFile(path), caption=f"📦 XLSX отчёт по текущей смене {active['id']} готов.", reply_markup=stats_menu_kb())


@dp.message(F.text == "📦 XLSX смены")
async def selected_shift_xlsx(message: Message, state: FSMContext):
    data = await state.get_data()
    shift_id = data.get("selected_shift_id")
    if not shift_id:
        await message.answer("⚠️ Сначала выбери смену через 📋 Список смен.", reply_markup=stats_menu_kb())
        return
    path = export_shift_to_xlsx(message.from_user.id, shift_id)
    if not path:
        await message.answer("📭 Нет данных для экспорта.", reply_markup=selected_shift_kb())
        return
    await message.answer_document(FSInputFile(path), caption=f"📦 XLSX отчёт по смене {shift_id} готов.", reply_markup=selected_shift_kb())


@dp.message(F.text == "📌 Ближайшие матчи")
async def upcoming_matches_handler(message: Message):
    rows = all_rows("""
        SELECT match_name, market, stake, odds, bookmaker, match_start_at
        FROM bets
        WHERE user_id=? AND result_status='pending' AND match_start_at IS NOT NULL
        ORDER BY match_start_at ASC LIMIT 20
    """, (message.from_user.id,))
    future = []
    for r in rows:
        try:
            dt = datetime.fromisoformat(r["match_start_at"])
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=TIMEZONE)
        except Exception:
            continue
        if dt >= now_dt():
            future.append((r, dt))
    if not future:
        await message.answer("📭 Ближайших pending матчей нет.", reply_markup=stats_menu_kb())
        return
    lines = ["📌 <b>Ближайшие матчи</b>\n"]
    for i, (r, dt) in enumerate(future[:10], 1):
        lines.append(
            f"{i}. <b>{r['match_name']}</b>\n"
            f"📌 {r['market']}\n"
            f"💸 {r['stake']} | 📈 {r['odds']} | 🏦 {bookmaker_label(r['bookmaker'])}\n"
            f"🕒 {dt.strftime('%d.%m.%Y %H:%M')} МСК\n"
        )
    await message.answer("\n".join(lines), reply_markup=stats_menu_kb())


# =========================
# SERVICE
# =========================
@dp.message(F.text == "📋 Логи")
async def logs_handler(message: Message):
    rows = get_recent_logs(10)
    if not rows:
        await message.answer("📭 Логов нет.", reply_markup=service_menu_kb())
        return
    lines = ["📋 <b>Последние логи</b>\n"]
    for r in rows:
        lines.append(f"<b>{r['level']}</b> | {r['created_at']}\n{r['message']}\n")
    await message.answer("\n".join(lines), reply_markup=service_menu_kb())


# =========================
# UNIVERSAL TEXT / FORWARDED BET FLOW
# =========================
@dp.message(F.text)
async def universal_text_handler(message: Message, state: FSMContext):
    if message.from_user.id != OWNER_ID:
        return

    text = (message.text or "").strip()
    current_state = await state.get_state()

    ignored = {
        S.waiting_budget.state, S.waiting_end_shift_confirm.state,
        S.waiting_shift_number.state, S.waiting_shift_bet_number.state,
        S.waiting_selected_bet_action.state, S.waiting_selected_bet_new_stake.state,
        S.waiting_selected_bet_result_status.state, S.waiting_risk_decision.state,
    }
    if current_state in ignored:
        return

    if current_state == S.waiting_bet_amount.state:
        data = await state.get_data()
        pending = data.get("pending_bet")
        if not pending:
            await state.clear()
            await message.answer("⚠️ Не нашёл ожидаемую ставку.", reply_markup=bets_menu_kb())
            return
        try:
            amount = as_float(text)
            if amount <= 0:
                raise ValueError
        except Exception:
            await message.answer("⚠️ Сумма не распознана. Введи число, например: <code>1500</code>", reply_markup=amount_retry_kb())
            return
        active = get_active_shift(message.from_user.id)
        if not active:
            await state.clear()
            await message.answer("📭 Активной смены нет.", reply_markup=shift_menu_kb(False))
            return
        try:
            add_bet_db(active["id"], message.from_user.id, pending, amount)
        except Exception as e:
            if "UNIQUE constraint failed" in str(e):
                await state.clear()
                await message.answer("⚠️ Эта ставка уже была добавлена ранее.", reply_markup=bets_menu_kb())
                return
            log_error(f"Bet insert failed: {e}")
            await state.clear()
            await message.answer(f"❌ Ошибка записи ставки: <code>{e}</code>", reply_markup=bets_menu_kb())
            return

        new_spent = round(active["spent"] + amount, 2)
        remain = round(active["budget"] - new_spent, 2)
        warn = f"\n\n⚠️ <b>Выход за лимит</b> на <b>{round(new_spent - active['budget'], 2)}</b>" if new_spent > active["budget"] else ""
        await state.clear()
        await message.answer(
            f"✅ <b>Ставка сохранена</b>\n\n"
            f"💸 Сумма: <b>{amount}</b>\n"
            f"📊 Поставлено: <b>{new_spent}</b> / <b>{active['budget']}</b>\n"
            f"🟢 Остаток: <b>{remain}</b>{warn}",
            reply_markup=bets_menu_kb(),
        )
        return

    if not is_forward_message(message):
        return

    if not get_active_shift(message.from_user.id):
        await message.answer("⚠️ Сначала начни смену.", reply_markup=shift_menu_kb(False))
        return

    parsed = parse_bet(text)
    if not parsed:
        diagnostics = parse_bet_diagnostics(text)
        await message.answer(
            "⚠️ <b>Ставка не распознана</b>\n\n"
            "<b>Что не получилось определить:</b>\n"
            f"{diagnostics}\n\n"
            "<b>Что сделать:</b>\n"
            "Перешли исходное сообщение без ручного редактирования.",
            reply_markup=bets_menu_kb(),
        )
        log_warning(f"Bet parse failed | diagnostics: {diagnostics}")
        return

    parsed = analyze_risk(parsed)
    await state.update_data(pending_bet=parsed)
    await state.set_state(S.waiting_risk_decision)

    match_start = datetime.fromisoformat(parsed["match_start_at"]).astimezone(TIMEZONE).strftime("%d.%m.%Y %H:%M")
    risk_text = compact_risk_text(parsed)

    await message.answer(
        "🎯 <b>СТАВКА РАСПОЗНАНА</b>\n━━━━━━━━━━━━━━\n"
        f"🏅 <b>{parsed['sport']}</b>\n"
        f"🏆 {parsed['tournament'] or '-'}\n\n"
        f"🏟 <b>{parsed['match_name']}</b>\n\n"
        f"📌 {parsed['market']}\n"
        f"📊 Группа: <b>{parsed.get('market_group')}</b>\n\n"
        f"📈 КФ: <b>{parsed['odds']}</b>\n"
        f"🧠 EV: <b>{parsed['ev'] if parsed['ev'] is not None else '-'}</b>\n"
        f"🏦 {bookmaker_label(parsed['bookmaker'])}\n"
        f"🕒 Старт: <b>{match_start} МСК</b>\n\n"
        f"{risk_text}\n\n"
        f"<b>Что делаем?</b>",
        reply_markup=risk_decision_kb(),
    )


# =========================
# REMINDERS
# =========================
async def reminder_job():
    for r in get_due_reminders():
        dt_text = r["dt"].astimezone(TIMEZONE).strftime("%d.%m.%Y %H:%M")
        try:
            await bot.send_message(
                r["user_id"],
                "⏰ <b>Напоминание</b>\n\n"
                f"Через {REMINDER_MINUTES} минут матч:\n"
                f"🏟 <b>{r['match_name']}</b>\n"
                f"📌 {r['market']}\n"
                f"💸 {r['stake']}\n"
                f"📈 {r['odds']}\n"
                f"🏦 {bookmaker_label(r['bookmaker'])}\n"
                f"🕒 {dt_text} МСК",
                reply_markup=bets_menu_kb(),
            )
            mark_reminder_sent(r["id"])
            log_info(f"Reminder sent | bet_id={r['id']}")
        except Exception as e:
            log_error(f"Reminder failed | bet_id={r['id']} | error={e}")


# =========================
# STARTUP
# =========================
async def main():
    print("BOT STARTED")
    init_db()
    scheduler.add_job(reminder_job, "interval", seconds=30, max_instances=1, coalesce=True)
    scheduler.start()
    log_info("Bot started")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
